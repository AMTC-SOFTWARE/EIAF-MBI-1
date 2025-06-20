from PyQt5.QtCore import QObject, QState, pyqtSlot, pyqtSignal
from datetime import datetime
from paho.mqtt import publish
from threading import Timer
########### MODIFICACION ########### 
from time import sleep
from openpyxl import Workbook, load_workbook #crear un excel
from openpyxl.utils import get_column_letter #obtener columna de excel
from time import strftime #para obtener hora actual y fecha actual gg
########### MODIFICACION ########### 
from os.path import exists
from time import strftime
from pickle import TRUE, load
from cv2 import imread
from os import system
from copy import copy
import requests
import pprint
import json
import cv2
import pandas as pd
from admin import Admin       

class Startup(QState):
    ok  = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        Timer(0.05, self.model.log, args = ("STARTUP",)).start()
        if self.model.local_data["user"]["type"] != "":
            Timer(0.05, self.logout, args = (copy(self.model.local_data["user"]),)).start()
        command = {
            "popout_relay" : {"text": "close", "color": "black"},
            "lbl_info0" : {"text": "", "color": "black"},
            "lbl_info1" : {"text": "", "color": "black"},
            #"lbl_info2" : {"text": "", "color": "green"}, #debe ir comentado para evitar que se reinicie el mensaje de fusibles que faltan por rellenar
            "lbl_info3" : {"text": "", "color": "black"},
            "lbl_info4" : {"text": "", "color": "black"},
            "lbl_nuts" :  {"text": "", "color": "purple"},
            #"lbl_nuts" : {"text": "  F1: Enviar a Home\nF12: Reiniciar Robots", "color": "purple"},
            ##############################################
            "lbl_box1"  : {"text": "", "color": "black"},
            "lbl_box2"  : {"text": "", "color": "black"},
            "lbl_box3"  : {"text": "", "color": "black"},
            "lbl_box4"  : {"text": "", "color": "black"},
            "lbl_box5"  : {"text": "", "color": "black"},
            "lbl_box6"  : {"text": "", "color": "black"},
            "lbl_box7"  : {"text": "", "color": "black"}, ######### Modificación para F96 #########
            "lbl_box8" :  {"text": "", "color": "orange"},
            "lbl_box9" :  {"text": "", "color": "orange"},
            "lbl_box10" : {"text": "", "color": "orange"},
            "lbl_box11" : {"text": "", "color": "orange"},
            "lbl_box12" : {"text": "", "color": "orange"},
            "lbl_box13" : {"text": "", "color": "orange"},
            ##############################################
            "lbl_result" : {"text": "Se requiere un login para continuar", "color": "green"},
            "lbl_steps" : {"text": "Ingresa tu código de acceso", "color": "black"},
            "lbl_user" : {"type":"", "user": "", "color": "black"},
            "img_user" : "blanco.jpg",
            "img_nuts" : "blanco.jpg",
            "img_center" : "logo.jpg",
            "img_fuse": "vacio2.jpg",
            "show":{"scanner": False}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        #self.hideSoftware()
        #Timer(1, self.hideSoftware).start()
        ##############################################Timer(1, self.kioskMode).start()
        try:
            turnos = {
            "1":["07-00","18-59"],
            "2":["19-00","06-59"],
            }
            endpoint = "http://{}/contar/historial/FIN".format(self.model.server)
            response = requests.get(endpoint, data=json.dumps(turnos))
            response = response.json()
            print("response: ",response)
            print("Startup para mostrar conteo de arneses")
            command["lcdNumber"] = {"value": response["conteo"]}
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        except Exception as ex:
            print("Error en el conteo ", ex)
        if exists("data\config"):
            with open("data\config", "rb") as f:
                data = load(f)
                self.model.config_data.update(data)
        print("\nConfig:\n", self.model.config_data, "\n")

        ##############
        ##############
        self.ok.emit()
        ##############
        ##############

    def kioskMode(self):
        system("taskkill /f /im explorer.exe")

    def hideSoftware(self):
        publish.single("System",json.dumps({"window" : False}),hostname='127.0.0.1', qos = 2)
        #publish.single("visycam/set",json.dumps({"window" : False}),hostname='127.0.0.1', qos = 2)

    def logout(self, user):
        try:
            Timer(0.05, self.model.log, args = ("LOGOUT",)).start() 
            data = {
                "NOMBRE": self.model.local_data["user"]["name"],
                "GAFET": self.model.local_data["user"]["pass"],
                "TIPO": self.model.local_data["user"]["type"],
                "SESION": "LOGOUT",
                "FECHA": datetime.now().isoformat()
                }
            resp = requests.post(f"http://{self.model.server}/api/post/manager", data=json.dumps(data))
        except Exception as ex:
            print("Logout Exception: ", ex)
        finally:
            self.model.local_data["user"]["type"] = ""
            self.model.local_data["user"]["name"] = ""
            self.model.local_data["user"]["pass"] = ""

class Login (QState):
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model
    def onEntry(self, event):
        command = {
            "show":{"login": True},
            "allow_close": True
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class CheckLogin (QState):
    ok      = pyqtSignal()
    nok     = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        command = {
            "lbl_result" : {"text": "ID recibido", "color": "green"},
            "lbl_steps" : {"text": "Validando usuario...", "color": "black"},
            "show":{"login": False}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        Timer(0.05,self.API_requests).start()

    def API_requests (self):
        try:
            endpoint = ("http://{}/api/get/usuarios/GAFET/=/{}/ACTIVO/=/1".format(self.model.server, self.model.gui["ID"]))
            response = requests.get(endpoint).json()
    
            if "ACTIVO" in response and response["ACTIVO"]:
                self.model.local_data["user"]["type"] = response["TIPO"][0]
                self.model.local_data["user"]["name"] = response["NOMBRE"][0]
                self.model.local_data["user"]["pass"] = copy(self.model.gui["ID"][0])
                data = {
                    "NOMBRE": self.model.local_data["user"]["name"],
                    "GAFET": self.model.local_data["user"]["pass"],
                    "TIPO": self.model.local_data["user"]["type"],
                    "SESION": "LOGIN",
                    "FECHA": datetime.now().isoformat()
                    }
                resp = requests.post(f"http://{self.model.server}/api/post/manager", data=json.dumps(data))
                command = {
                    "lbl_user" : {
                        "type":self.model.local_data["user"]["type"],
                        "user": self.model.local_data["user"]["name"], 
                        "color": "black"
                        },
                    "img_user" : self.model.local_data["user"]["name"] + ".jpg"
                    }
                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                Timer(0.05, self.model.log, args = ("LOGIN",)).start()
                self.ok.emit()
            else:
                 command = {
                    "lbl_result" : {"text": "Intentalo de nuevo", "color": "green"},
                    "lbl_steps" : {"text": "Ingresa tu código de acceso", "color": "black"}
                    }
                 publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                 self.nok.emit()
        except Exception as ex:
            print("Login request exception: ", ex)
            command = {
                "lbl_result" : {"text": "Intentalo de nuevo", "color": "red"},
                "lbl_steps" : {"text": "Ingresa tu código de acceso", "color": "black"}
                }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            self.nok.emit()

class StartCycle (QState):
    ok = pyqtSignal()
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        #limpiar variable para caja anterior que pasó
        self.model.reset()

        #se reinicia variable que dice que el robot A finalizó
        self.model.robot_a_terminado = False
        #se reinician variables por si se da llave en el momento de la inserción manual del relay
        self.model.waiting_button_inserted_singal["robot_a"] = False
        self.model.waiting_button_inserted_singal["robot_b"] = False
        self.model.conjunto = []

        if self.model.modo_manual_activado==False:
            Timer(1, self.robots_home).start()
        Timer(0.05, self.model.log, args = ("IDLE",)).start() 
        command = {
            "popout_relay" : {"text": "close", "color": "red"},
            "lbl_info1" : {"text": "", "color": "black"},
            #"lbl_info2" : {"text": "", "color": "green"},
            "lbl_info3" : {"text": "", "color": "black"},
            "lbl_nuts" : {"text": "", "color": "purple"},
            #"lbl_nuts" : {"text": "  F1: Enviar a Home\nF12: Reiniciar Robots", "color": "purple"},
            #############################################
            "lbl_box1" : {"text": "", "color": "orange"},
            "lbl_box2" : {"text": "", "color": "orange"},
            "lbl_box3" : {"text": "", "color": "orange"},
            "lbl_box4" : {"text": "", "color": "orange"},
            "lbl_box5" : {"text": "", "color": "orange"},
            "lbl_box6" : {"text": "", "color": "orange"},
            "lbl_box7" : {"text": "", "color": "orange"},######### Modificación para F96 #########
            "lbl_box8" : {"text": "", "color": "orange"},
            "lbl_box9" : {"text": "", "color": "orange"},
            "lbl_box10" : {"text": "", "color": "orange"},
            "lbl_box11" : {"text": "", "color": "orange"},
            "lbl_box12" : {"text": "", "color": "orange"},
            "lbl_box13" : {"text": "", "color": "orange"},
            #############################################
            "lbl_result" : {"text": "Nuevo ciclo, Coloca las cajas. Seleccionar:", "color": "green"},
            "lbl_steps" : {"text": '"CTRL" / "START" para DOS Robots; "F4" para UN Robot', "color": "black"},
            "img_nuts" : "blanco.jpg",
            "img_center" : "logo.jpg",
            "img_fuse": "vacio2.jpg",
            "allow_close": False,
            "cycle_started": False,
            "statusBar": "clear"
            }
        if self.model.shutdown == True:
            Timer(0.05, self.logout, args = (self.model.local_data["user"],)).start()
            command["lbl_result"] = {"text": "Apagando equipo...", "color": "green"}
            command["lbl_steps"] = {"text": ""}
            command["shutdown"] = True
            Timer(3, self.clamps_release).start()
        if self.model.config_data["trazabilidad"] == True:
            command["lbl_info3"] = {"text": "Trazabilidad\nActivada", "color": "green"}
        else:
            command["lbl_info3"] = {"text": "Trazabilidad\nDesactivada", "color": "red"}
        if self.model.config_data["modo_manual"] == True:
            command["lbl_info0"] = {"text": "MODO\nMANUAL", "color": "darkmagenta"}
            command["lbl_result"] = {"text": "Nuevo Ciclo", "color": "green"}
            command["lbl_steps"] = {"text": 'Presionar "CTRL" para Comenzar ciclo MANUAL', "color": "black"}
        else:
            command["lbl_info0"] = {"text": "", "color": "red"}
        command["lcdNumber"] = {"value": 0, "visible": True}
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        try:
            turnos = {
            "1":["07-00","18-59"],
            "2":["19-00","06-59"],
            }
            endpoint = "http://{}/contar/historial/FIN".format(self.model.server)
            response = requests.get(endpoint, data=json.dumps(turnos))
            response = response.json()
            print("response: ",response)
            print("Startup para mostrar conteo de arneses")
            command["lcdNumber"] = {"value": response["conteo"]}
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        except Exception as ex:
            print("Error en el conteo ", ex)
        try:
            turnos = {
            "1":["07-00","18-59"],
            "2":["19-00","06-59"],
            }
            horario_turno1={"7":0,
                        "8":0,
                        "9":0,
                        "10":0,
                        "11":0,
                        "12":0,
                        "13":0,
                        "14":0,
                        "15":0,
                        "16":0,
                        "17":0,
                        "18":0,
                        "19":0,
                        "20":0,
                        "21":0,
                        "22":0,
                        "23":0,
                        "00":0,
                        "01":0,
                        "02":0,
                        "03":0,
                        "04":0,
                        "05":0,
                        "06":0,
                        }
            endpoint = "http://{}/contar/historial/FIN".format(self.model.server)
            response = requests.get(endpoint, data=json.dumps(turnos))
            response = response.json()
            print("response: ",response)
            print("Startup para mostrar conteo de arneses")
            command["lcdNumber"] = {"value": response["conteo"]}
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)


            endpoint = "http://{}/horaxhora/historial/FIN".format(self.model.server)
            response = requests.get(endpoint, data=json.dumps(turnos))
            response = response.json()
            
            arneses_turno=pd.DataFrame({'HM': response['HM'],
                   'INICIO': response['INICIO'],
                   'FIN': response['FIN'],
                   'RESULTADO': response['RESULTADO'],
                   'USUARIO': response['USUARIO']})
            
            
            arneses_turno['INICIO']=pd.to_datetime(arneses_turno['INICIO'])
            arneses_turno['FIN']=pd.to_datetime(arneses_turno['FIN'])
            arneses_turno['RESULTADO']=arneses_turno['RESULTADO'].astype("string")

            base_temporal = arneses_turno[(arneses_turno["RESULTADO"]=="BUENO")]
            #Calcula Duración de ciclo de los arneses
            arneses_turno["INTERVALO"]=base_temporal['FIN']-base_temporal['INICIO']
            
            promedio_ciclo_turno=arneses_turno["INTERVALO"].mean().total_seconds() / 60
            
            # Obtener la parte entera y decimal
            parte_entera = int(promedio_ciclo_turno)
            parte_decimal = promedio_ciclo_turno - parte_entera
            
            # Convertir la parte decimal a segundos
            segundos = round(parte_decimal * 60)
            if segundos<10:
                segundos="0"+str(segundos)
            tiempo_ciclo_promedio=str(parte_entera)+":"+str(segundos)

            command = {
            "lcdNumtiempo": {"label_name": "Tiempo Ciclo\n Promedio", "color":"#68FD94", "value": tiempo_ciclo_promedio}
            }


            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)


        except Exception as ex:
            print("Error en el promedio ", ex)
      
    def robots_home (self):
        print("self.model.modo_manual_activado",self.model.modo_manual_activado)
        if self.model.modo_manual_activado==False:
            ########### MODIFICACION ###########
            self.model.robothome_a = True # variable para activar Mensaje de enviar robot a home, se resetea sola en comm.py
            self.model.robothome_b = True # variable para activar Mensaje de enviar robot a home, se resetea sola en comm.py
            #Enviar 2 stop, para asegurar y después con un start el robot debe ir solo a Home
            publish.single(self.model.pub_topics["robot_a"],json.dumps({"command": "stop"}),hostname='127.0.0.1', qos = 2)
            sleep(0.3)
            publish.single(self.model.pub_topics["robot_b"],json.dumps({"command": "stop"}),hostname='127.0.0.1', qos = 2)
            sleep(0.3)
            publish.single(self.model.pub_topics["robot_a"],json.dumps({"command": "stop"}),hostname='127.0.0.1', qos = 2)
            sleep(0.3)
            publish.single(self.model.pub_topics["robot_b"],json.dumps({"command": "stop"}),hostname='127.0.0.1', qos = 2)
            sleep(0.3)
            publish.single(self.model.pub_topics["robot_a"],json.dumps({"command": "start"}),hostname='127.0.0.1', qos = 2)
            sleep(0.3)
            publish.single(self.model.pub_topics["robot_b"],json.dumps({"command": "start"}),hostname='127.0.0.1', qos = 2)
            ########### MODIFICACION ###########
            #publish.single(self.model.pub_topics["robot_a"],json.dumps({"trigger": "HOME"}),hostname='127.0.0.1', qos = 2)
            #publish.single(self.model.pub_topics["robot_b"],json.dumps({"trigger": "HOME"}),hostname='127.0.0.1', qos = 2)

    def clamps_release(self):
        command = {}
        for i in self.model.fuses_BB:
            command[i] = False
        command["ERROR_insertion"] = False
        publish.single(self.model.pub_topics["plc"],json.dumps(command),hostname='127.0.0.1', qos = 2)

    def logout(self, user):
        try:
            Timer(0.05, self.model.log, args = ("LOGOUT",)).start() 
            data = {
                "NOMBRE": self.model.local_data["user"]["name"],
                "GAFET": self.model.local_data["user"]["pass"],
                "TIPO": self.model.local_data["user"]["type"],
                "SESION": "LOGOUT",
                "FECHA": datetime.now().isoformat()
                }
            resp = requests.post(f"http://{self.model.server}/api/post/manager", data=json.dumps(data))
        except Exception as ex:
            print("Logout Exception: ", ex)

    def onExit(self, QEvent):
        command = {
            "lbl_result" : {"text": "Nidos activados", "color": "green"},
            "lbl_steps" : {"text": "Escanea el DATAMATRIX", "color": "black"},
            "cycle_started": True
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class Config (QState):
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model
        self.admin = None

    def onEntry(self, event):
        Timer(0.05, self.model.log, args = ("CONFIG",)).start() 
        admin = Admin(data = self.model)
        command = {
            "lbl_result" : {"text": "Sistema en configuración", "color": "green"},
            "lbl_steps" : {"text": "Ciclo de operación deshabilitado", "color": "black"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class ScanQr (QState):
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        command = {
            "show":{"scanner": True}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        #Se comenta para editar el clampeo antes de identificar el arnés a trabajar.
        #command = {}
        #for i in self.model.fuses_BB:
        #    if "PDC-R" in i:
        #        command[i] = False
        #    else:
        #        command[i] = True
        #publish.single(self.model.pub_topics["plc"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        Timer(0.05, self.model.log, args = ("RUNNING",)).start()

    def onExit(self, QEvent):
        command = {
            "show":{"scanner": False}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class StandbyTraza (QState):
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        print("Estado: StandbyTraza, esperando señal de ctrl o start para continuar")

class CheckQr (QState):

    ok_F4   = pyqtSignal()
    ok_CTRL   = pyqtSignal()
    ok_MANUAL = pyqtSignal() 

    nok     = pyqtSignal()
    rework  = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        command = {
            "lbl_result" : {"text": "Datamatrix escaneado", "color": "green"},
            "lbl_steps" : {"text": "Validando", "color": "black"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        Timer(0.05, self.API_requests).start()

    def API_requests (self):
        try:
            print("Estado de Sistema de Trazabilidad: ",self.model.config_data["trazabilidad"])
            orden = None
            dbEvent = None
            coincidencias = 0
            self.model.codes["FET"] = self.model.gui["code"]
            temp = self.model.gui["code"].split (" ")
            self.model.codes["HM"] = "--"
            self.model.codes["REF"] = "--"
            #correct_lbl = False
            correct_lbl = True
            self.model.cronometro_ciclo=True
            for i in temp:
                if "HM" in i:
                    self.model.codes["HM"] = i
                    
                    if "HM000000011936" in i:
                        self.model.config_data["trazabilidad"] = False
                        
                    if "HM000000011925" in i:
                        self.model.config_data["trazabilidad"] = False

                    if "HM000000011920" in i:
                        self.model.config_data["trazabilidad"] = False


                    if self.model.config_data["trazabilidad"] == False:
                        command = {
                            "lbl_info3" : {"text": "Trazabilidad\nDesactivada", "color": "red"}
                        }
                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)


                if "IL" in i or "IR" in i:
                    self.model.codes["REF"] = i
                if "EL." in i:
                    correct_lbl = True
            if not(correct_lbl):
                self.model.cronometro_ciclo=False
                command = {
                        "lbl_result" : {"text": "Datamatrix incorrecto", "color": "red"},
                        "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                        }
                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                self.nok.emit()
                return

            #### Trazabilidad FAMX2
            if self.model.config_data["trazabilidad"]:
                try:
                    print("||||||||||||Consulta de HM a FAMX2...")
                    endpoint = "http://{}/seghm/get/seghm/NAMEPREENSAMBLE/=/INTERIOR/HM/=/{}".format(self.model.server,self.model.codes["HM"])
                    famx2response = requests.get(endpoint).json()
                    print("Respuesta de FAMX2: \n",famx2response)
                    #No existen coincidencias del HM en FAMX2
                    if "items" in famx2response:
                        self.model.cronometro_ciclo=False
                        print("ITEMS por que no se encontraron coincidencias en FAMX2")
                        command = {
                            "lbl_result" : {"text": "HM no registrado en Sistema de Trazabilidad", "color": "red"},
                            "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                            }
                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                        self.nok.emit()
                        return
                    #Si existe el HM en FAMX2
                    else:
                        print("FAMX2 Salida de Torque: \n",famx2response["SALTORQUE"])
                        print("FAMX2 Ubicación: \n",famx2response["UBICACION"])
                        famx2response["UBICACION"] = famx2response["UBICACION"].replace(" ","")
                        print("FAMX2 Ubicación sin espacios: \n",famx2response["UBICACION"])

                        #Si la columna que indica la hora de salida de TORQUE, es diferente a None, significa que completó esa estación y SI puede entrar a Inserción.
                        if famx2response["SALTORQUE"] != None: #AQUIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIII
                            print("El arnés ya salió de TORQUE")
                            #La ubicación debe ser "SALIDA_DE_TORQUE"
                            if famx2response["UBICACION"] == "SALIDA_DE_TORQUE" or famx2response["UBICACION"] == "ENTRADA_A_INSERCION":

                                #Se guarda el id del arnés de FAMX2 en el modelo para realizar updates en el servidor de FAMX2.
                                self.model.id_HM = famx2response["id"]
                                self.model.datetime = datetime.now()
                                #### Trazabilidad FAMX2 Update de Información
                                print("||Realizando el Update de ENTRADA a Trazabilidad en FAMX2")
                                print("ID a la que se realizará el Update para Trazabilidad",self.model.id_HM)

                                entTrazabilidad = {
                                    "ENTINSERCION": self.model.datetime.strftime("%Y/%m/%d %H:%M:%S"),
                                    "UBICACION": "ENTRADA_A_INSERCION",
                                    "NAMEINSERCION": self.model.no_serie
                                    }

                                endpointUpdate = "http://{}/seghm/update/seghm/{}".format(self.model.server,self.model.id_HM)
                                respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(entTrazabilidad))
                                respTrazabilidad = respTrazabilidad.json()
                                print("respTrazabilidad del update: ",respTrazabilidad)

                                if "exception" in respTrazabilidad:
                                    sleep(0.5)
                                    respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(entTrazabilidad))
                                    respTrazabilidad = respTrazabilidad.json()
                                    print("respTrazabilidad del update: ",respTrazabilidad)

                                    if "exception" in respTrazabilidad:
                                        sleep(0.5)
                                        respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(entTrazabilidad))
                                        respTrazabilidad = respTrazabilidad.json()
                                        print("respTrazabilidad del update: ",respTrazabilidad)

                                        if "exception" in respTrazabilidad:
                                            self.model.cronometro_ciclo=False
                                            print("no se logró hacer el update en trazabilidad")
                                            command = {
                                                        "lbl_result" : {"text": "No se logró hacer el update de Trazabilidad", "color": "red"},
                                                        "lbl_steps" : {"text": "Intenta Ingresar nuevamente arnés", "color": "black"}
                                                        }
                                            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                                            self.nok.emit()
                                            return
                                #### Trazabilidad FAMX2 Update de Información

                            else:                    
                                self.model.cronometro_ciclo=False
                                print("El Arnés se encuentra en otra ubicación de entrada")
                                command = {
                                "lbl_result" : {"text": "Ubicación de HM Incorrecta:", "color": "red"},
                                "lbl_steps" : {"text": famx2response["UBICACION"], "color": "black"}
                                }
                                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                                self.nok.emit()
                                return

                        #Si la columna que indica la hora de salida de TORQUE es None, significa que no ha completado esa estación y NO puede entrar aún a Inserción.
                        else:
                            self.model.cronometro_ciclo=False
                            print("El Arnés no ha pasado por la estación anterior (TORQUE) por lo que no puede entrar a Torque")
                            command = {
                            "lbl_result" : {"text": "Arnés sin Fecha de Historial de TORQUE", "color": "red"},
                            "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                            }
                            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                            self.nok.emit()
                            return
                except Exception as ex:
                    self.model.cronometro_ciclo=False
                    print("Conexión con FAMX2 exception: ", ex)
                    command = {
                            "lbl_result" : {"text": "Error de Conexión con Sistema de Trazabilidad", "color": "red", "font": "40pt"},
                            "lbl_steps" : {"text": "Verifique su conexión o deshabilite el Sistema de Trazabilidad con supervisión \nde personal de calidad", "color": "black", "font": "22pt"}
                            }
                    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                    self.nok.emit()
                    return
            ####

            endpoint = "http://{}/api/get/eventos".format(self.model.server)
            eventos = requests.get(endpoint).json()
            print("Lista eventos:\n",eventos)
            #print("Eventos: ",eventos["eventos"])
            #print("Eventos KEYS: ",eventos["eventos"].keys())
            for key in eventos["eventos"].keys():
                print("++++++++++++++Evento Actual++++++++++++++++:\n ",key)
                print("Valor Activo del Evento actual: ",eventos["eventos"][key][1][0])
                if eventos["eventos"][key][1][0] == 1:
                    endpoint = "http://{}/api/get/{}/modularidades/MODULARIDAD/=/{}/ACTIVO/=/1".format(self.model.server, key, self.model.codes["REF"])
                    response = requests.get(endpoint).json()
                    #print("Response: ",response)
                    if "MODULARIDAD" in response:
                        for i in response:
                            response[i] = response[i][0]
                        dbEvent = key
                        coincidencias += 1
                        print("En este Evento se encuentra la modularidad \n")
                        orden = response
            print("Coincidencias = ",coincidencias)
            if dbEvent != None:
                print("La Modularidad pertenece al Evento: ",dbEvent)
                if coincidencias != 1:
                    self.model.cronometro_ciclo=False
                    print("Datamatrix Redundante")
                    command = {
                        "lbl_result" : {"text": "Datamatrix redundante", "color": "red"},
                        "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                        }
                    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                    self.nok.emit()
                    return
                else:
                    print("Datamatrix Correcto")
            else:
                self.model.cronometro_ciclo=False
                print("La Modularidad NO pertenece a ningún evento")
                command = {
                    "lbl_result" : {"text": "Datamatrix no registrado", "color": "green"},
                    "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                    }
                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                self.nok.emit()
                return

            #endpoint = "http://{}/api/get/modularidades/MODULARIDAD/=/{}/ACTIVO/=/1".format(self.model.server, self.model.codes["REF"])
            #response = requests.get(endpoint).json()

            #if "MODULARIDAD" in response:
            #    if len(response["MODULARIDAD"]) == 1: 
            #        for i in response:
            #            response[i] = response[i][0]
            #        if response["ACTIVO"]:
            #            orden = response
            #        else:
            #            command = {
            #                        "lbl_result" : {"text": "Datamatrix desactivado", "color": "red"},
            #                        "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
            #                      }
            #            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            #            self.nok.emit()
            #            return
            #    else: 
            #        command = {
            #                    "lbl_result" : {"text": "Datamatrix redundante", "color": "red"},
            #                    "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
            #                  }
            #        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            #        self.nok.emit()
            #        return
            #else:
            #    command = {
            #            "lbl_result" : {"text": "Datamatrix no registrado", "color": "red"},
            #            "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
            #            }
            #    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            #    self.nok.emit()
            #    return

            endpoint = "http://{}/api/get/{}/pdcr/variantes".format(self.model.server, dbEvent)
            pdcrVariantes = requests.get(endpoint).json()
            print("Lista Final de Variantes PDC-R:\n",pdcrVariantes)

            endpoint = "http://{}/api/get/historial/HM/=/{}/RESULTADO/=/BUENO".format(self.model.server, self.model.codes["HM"])
            response = requests.get(endpoint).json()

            if ("items" in response and not(response["items"])) or self.model.local_data["qr_rework"]:
                flag_s = False
                flag_m = False
                flag_l = False
                modules = orden["MODULOS_FUSIBLES"].split(sep = ",")
                print(f"\n\t\tMODULOS_FUSIBLES:\n{modules}")
                for s in pdcrVariantes["small"]:
                    if s in modules:
                        #print("Tiene un modulo de caja SMALL")
                        flag_s = True
                for m in pdcrVariantes["medium"]:
                    if m in modules:
                        #print("Tiene un modulo de caja Medium")
                        flag_m = True
                for l in pdcrVariantes["large"]:
                    if l in modules:
                        #print("Tiene un modulo de caja LARGE")
                        flag_l = True
                print("\t\tFLAGS:\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l)
                if flag_l == True:
                    variante = "PDC-R"
                if flag_m == True and flag_l == False:
                    variante = "PDC-RMID"
                if flag_s == True and flag_m == False:
                    variante = "PDC-RS"
                if flag_s == False and flag_m == False and flag_l == False:
                    variante = "N/A"
                    print("La caja no contiene módulos pertenecientes a las categorías.")
                
                print("MODULO DETERMINANTE: ",variante)

                for i in modules:

                    endpoint = "http://{}/api/get/{}/modulos_fusibles/MODULO/=/{}/_/=/_".format(self.model.server, dbEvent, i)
                    response = requests.get(endpoint).json()
                    if "MODULO" in response:
                        if len(response["MODULO"]) == 1: 
                            for j in response:
                                if j == "ID" or j == "MODULO":
                                    response[j] = response[j][0]
                                else:
                                    response[j] = json.loads(response[j][0])
                                    if j in self.model.database["fuses"]:
                                        if not len(response[j]):
                                            continue
                                        for k in response[j]:
                                            if response[j][k] == "empty":
                                                pass
                                            elif k in self.model.database["fuses"][j]:
                                                ########### MODIFICACION PARA DEFINIR LA VARIANTE DE CAJA PDC-R ###########
                                                if "PDC-R" in j:
                                                    #print("AQUI HAY UNA PDC-R: ",j)
                                                    #print("SU VALOR: ",response[j])
                                                    if flag_l:
                                                        #print("ESTA CAJA SE DEBE CONVERTIR A PDC-R LARGE")
                                                        boxVariant = "PDC-R"
                                                    if flag_m == True and flag_l == False:
                                                        #print("ESTA CAJA SE DEBE CONVERTIR A PDC-R MEDIUM")
                                                        boxVariant = "PDC-RMID"
                                                    if flag_s == True and flag_m == False:
                                                        #print("ESTA CAJA SE DEBE CONVERTIR A PDC-R SMALL")
                                                        boxVariant = "PDC-RMID"#Por el momento se colocó así para que cuando llegue una caja small se tome como si fuera mid
                                                    if flag_s == False and flag_m == False and flag_l == False:
                                                        print("ESTA MODULARIDAD NO CONTIENE MÓDULOS QUE DETERMINEN SU VARIANTE")
                                                        command = {
                                                                "lbl_result" : {"text": "Sin módulos que determinen su variante en PDC-R"},
                                                                "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                                                              }
                                                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                                                        self.nok.emit()
                                                        return
                                                    if not boxVariant in self.model.database["clamps"]:
                                                        self.model.database["clamps"].append(boxVariant)
                                                    if self.model.database["fuses"][boxVariant][k] == "empty":
                                                        self.model.database["fuses"][boxVariant][k] = response[j][k]
                                                    elif  self.model.database["fuses"][boxVariant][k] != response[j][k]:
                                                            command = {
                                                                "lbl_result" : {"text": f'DB Error con Módulo {response["MODULO"]}  en el fusible {boxVariant}: {k}", "color": "red'},
                                                                "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                                                              }
                                                            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                                                            self.nok.emit()
                                                            return
                                                else:
                                                    if not j in self.model.database["clamps"]:
                                                        self.model.database["clamps"].append(j)
                                                    if self.model.database["fuses"][j][k] == "empty":
                                                        self.model.database["fuses"][j][k] = response[j][k]
                                                    elif  self.model.database["fuses"][j][k] != response[j][k]:
                                                        self.model.cronometro_ciclo=False
                                                        command = {
                                                            "lbl_result" : {"text": f'DB Error con Módulo {response["MODULO"]}  en el fusible {j}: {k}", "color": "red'},
                                                            "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                                                          }
                                                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                                                        self.nok.emit()
                                                        return
                        else:
                            self.model.cronometro_ciclo=False
                            print("response[MODULO]",response["MODULO"])
                            command = {
                                    "lbl_result" : {"text": f"Módulo {i} redundante en Matriz de evento", "color": "red"},
                                    "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                                  }
                            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                            self.nok.emit()
                            return
                    else:
                        self.model.cronometro_ciclo=False
                        command = {
                                "lbl_result" : {"text": f"Modulo {i} no encontrado", "color": "red"},
                                "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                                }
                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                        self.nok.emit()
                        return 
                self.model.database["orden"] = orden
                self.model.datetime = datetime.now()

                self.model.QR = self.model.codes["HM"] + "-" + self.model.codes["REF"] + "-" + self.model.datetime.strftime("%d%m%Y%H%M%S") + "-" + self.model.no_serie
                print("\nQR: " + self.model.QR + "\n")
                print("\nAQUIIIII++++++: ",self.model.database["fuses"],"\n")

                #################### Distribucion de inserciones de acuerdo a cada robot de inserción ###########################
                temp = False
                for i in self.model.database["fuses"]:
                    for j in self.model.database["fuses"][i]:
                        if self.model.database["fuses"][i][j] != "empty":


                            if i == "PDC-RMID" and not(self.model.pdcr_mid):
                                self.model.database["fuses"]["PDC-R"] = {}
                                self.model.pdcr_mid = True
                            if i == "PDC-R" and not (temp):
                                self.model.database["fuses"]["PDC-RMID"] = {}
                                temp = True

                            ##################### SE ASIGNAN LAS TAREAS PARA EL ROBOTA y ROBOTB, EN LADO IZQ Y DER ####################

                            if i == "PDC-D" or i == "PDC-P" or i == "PDC-S17" or i == "PDC-S21":
                                #print("**** Robot A ****")
                                #print("Fusible: ",self.model.database["fuses"][i][j])
                                if self.model.database["fuses"][i][j] in self.model.AfusesIzq:
                                    #print("Fusible Robot A IZQUIERDA")
                                    self.model.robots["robot_a"]["queueIzq"].append([i, j, self.model.database["fuses"][i][j]])
                                if self.model.database["fuses"][i][j] in self.model.AfusesDer:
                                    #print("Fusible Robot A DERECHA")
                                    self.model.robots["robot_a"]["queueDer"].append([i, j, self.model.database["fuses"][i][j]])
                                

                            if i == "PDC-R" or i == "PDC-RMID" or i == "PDC-S" or i == "PDC-S9" or i == "PDC-S19" or i == "PDC-S20" or i == "F96-1" or i == "TBLU":
                                #print("**** Robot B ****")
                                #print("Fusible: ",self.model.database["fuses"][i][j])
                                if self.model.database["fuses"][i][j] in self.model.BfusesIzq:
                                    #print("Fusible Robot B IZQUIERDA")
                                    self.model.robots["robot_b"]["queueIzq"].append([i, j, self.model.database["fuses"][i][j]])
                                if self.model.database["fuses"][i][j] in self.model.BfusesDer:
                                    #print("Fusible Robot B DERECHA")
                                    self.model.robots["robot_b"]["queueDer"].append([i, j, self.model.database["fuses"][i][j]])
                #################################################################################################################         

                if self.model.local_data["qr_rework"]:
                    self.model.local_data["qr_rework"] = False
                print("dbEvent: ",dbEvent)
                event = dbEvent.upper()
                evento = event.replace('_',' ')

                if self.model.config_data["modo_manual"] == True:
                    command = {
                        "lbl_result" : {"text": "Datamatrix OK", "color": "green"},
                        "lbl_steps" : {"text": "MODO INSERCIÓN MANUAL", "color": "black"},
                        "statusBar" : orden["MODULARIDAD"]+" "+self.model.codes["HM"]+" "+evento,
                        #"img_center" : f"boxes/{batt}.jpg"  #Aqui actualizar la imagen principal con un colage de las cajas que faltan por clampear
                        }
                    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

                    self.model.modularity_manual.clear()

                    #SE GUARDA LA INFORMACION EN UNA SOLA LISTA
                    for elemento in self.model.robots["robot_a"]["queueIzq"]:
                        self.model.modularity_manual.append(elemento)
                    for elemento in self.model.robots["robot_a"]["queueDer"]:
                        self.model.modularity_manual.append(elemento)
                    for elemento in self.model.robots["robot_b"]["queueIzq"]:
                        if "F96" in elemento[1]:
                            elemento[0] = "F96"
                        self.model.modularity_manual.append(elemento)
                    for elemento in self.model.robots["robot_b"]["queueDer"]:
                        if "F96" in elemento[1]:
                            elemento[0] = "F96"
                        self.model.modularity_manual.append(elemento)

                    def ordenar_por_caja(elem):
                        cajas = {"PDC-R":1,"PDC-RMID":2,"F96":3,"PDC-S":4,"TBLU":5,"PDC-D":6,"PDC-P":7, "PDC-S19":8, "PDC-S20":9, "PDC-S17":10, "PDC-S21":11, "PDC-S9":12}
                        return cajas.get(elem[0],0)

                    self.model.modularity_manual = sorted(self.model.modularity_manual, key=ordenar_por_caja)
                    self.model.numero_de_cajas=len(self.model.modularity_manual[0])
                    self.model.modularity_manual_respaldo=copy(self.model.modularity_manual)
                    
                    for elemento in self.model.modularity_manual:
                        if not(elemento[0] in self.model.cajas_arnes):
                            self.model.cajas_arnes.append(elemento[0])

                    print("self.model.cajas_arnes",self.model.cajas_arnes)
                    print("self.model.cajas_arnes",len(self.model.cajas_arnes))
                    self.ok_MANUAL.emit()

                else:
                    command = {
                        "lbl_result" : {"text": "Datamatrix OK", "color": "green"},
                        "lbl_steps" : {"text": "Coloca el resto de las cajas", "color": "black"},
                        "statusBar" : orden["MODULARIDAD"]+" "+self.model.codes["HM"]+" "+evento,
                        #"img_center" : f"boxes/{batt}.jpg"  #Aqui actualizar la imagen principal con un colage de las cajas que faltan por clampear
                        }
                    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                    Timer(0.1, self.fuseBoxesClamps).start()

                    #################################
                    print("self.model.robots_mode EN CHECK_QR: ",self.model.robots_mode)
                    if self.model.robots_mode == 1:
                        #"MODO UN ROBOT ACTIVADO"
                        command = {"lbl_nuts" : {"text": "MODO: UN ROBOT", "color": "purple"}}
                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                        self.ok_F4.emit()

                    elif self.model.robots_mode == 2:
                        #"MODO DOS ROBOTS ACTIVADOS"
                        command = {"lbl_nuts" : {"text": "MODO: DOS ROBOTS", "color": "purple"}}
                        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                        self.ok_CTRL.emit()
                    else:
                        self.nok.emit()
                    ################################

            else:
                self.rework.emit()
                return

        except Exception as ex:
            print("Datamatrix request exception: ", ex) 
            temp = f"Database Exception: {ex.args}"
            command = {
                        "lbl_result" : {"text": temp, "color": "red"},
                        "lbl_steps" : {"text": "Inténtalo de nuevo", "color": "black"}
                        }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            self.model.fusesInit()
            self.nok.emit()

    def fuseBoxesClamps(self):
        command = {}
        for i in self.model.database["fuses"]:
            if i in self.model.database["clamps"]:
               if "PDC-S9" in i:
                    command[i] = False
               else:
                    command[i] = True
            else:
                command[i] = False

        print( f"cajas a clampear {command}")
        publish.single(self.model.pub_topics["plc"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class QrRework (QState):
    ok = pyqtSignal()
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

        self.model.transitions.key.connect(self.rework)
        self.model.transitions.code.connect(self.noRework)

    def onEntry(self, QEvent):
        command = {
            "lbl_result" : {"text": "Datamatrix procesado anteriormente", "color": "red"},
            "lbl_steps" : {"text": "Escanea otro código o gira la llave para continuar", "color": "black"},
            "show":{"scanner": True}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

    def onExit(self, QEvent):
        command = {
            "show":{"scanner": False}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

    def rework (self):
        self.model.local_data["qr_rework"] = True
        Timer(0.05, self.ok.emit).start()

    def noRework(self):
        Timer(0.05, self.ok.emit).start()

class ClampsMonitor(QState):
    ok = pyqtSignal()

    def __init__(self, module = "clamps_a", model = None, parent = None):
        super().__init__(parent)
        self.model = model
        self.module = module

    def onEntry(self, QEvent):

        temp = False
        #self.model.database["clamps"] contiene los clamps necesarios para el arnés escaneado
        #self.model.plc["clamps"] contiene las cajas que se han clampeado correctamente en físico
        print("\n database: ", self.model.database["clamps"])
        print(" PLC     : ", self.model.plc["clamps"],"\n")
        
        database_temp = []

        command = {
                    "lbl_result" : {"text": "Esperando cajas para continuar", "color": "green"},
                    "lbl_steps" : {"text": "Coloca el resto de las cajas", "color": "black"},
                    "img_center" : "logo.jpg"
                    }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        if self.module == "clamps_a":
            if not("PDC-D" in database_temp):
                if "PDC-D" in self.model.database["clamps"]:
                     database_temp.append("PDC-D")
            if not("PDC-S17" in database_temp):
                if "PDC-S17" in self.model.database["clamps"]:
                    database_temp.append("PDC-S17")
            if not("PDC-P" in database_temp):
                if "PDC-P" in self.model.database["clamps"]:
                    database_temp.append("PDC-P")

        if self.module == "clamps_b":
            if not("PDC-R" in  database_temp):
                if "PDC-R" in self.model.database["clamps"]:
                     database_temp.append("PDC-R")
            if not("PDC-RMID" in  database_temp):
                if "PDC-RMID" in self.model.database["clamps"]:
                     database_temp.append("PDC-RMID")
            if not("PDC-S" in  database_temp):
                if "PDC-S" in self.model.database["clamps"]:
                     database_temp.append("PDC-S")
            if not("PDC-S9" in  database_temp):
                if "PDC-S9" in self.model.database["clamps"]:
                    self.model.pdcs9_flag = True
                    database_temp.append("PDC-S9")
            if not("PDC-S19" in  database_temp):
                if "PDC-S19" in self.model.database["clamps"]:
                     database_temp.append("PDC-S19")
            if not("PDC-S20" in  database_temp):
                if "PDC-S20" in self.model.database["clamps"]:
                     database_temp.append("PDC-S20")
            if not("PDC-S21" in  database_temp):
                if "PDC-S21" in self.model.database["clamps"]:
                    database_temp.append("PDC-S21")
            if not("F96-1" in  database_temp):
                if "F96-1" in self.model.database["clamps"]:
                     database_temp.append("F96-1")
            if not("TBLU" in  database_temp):
                if "TBLU" in self.model.database["clamps"]:
                     database_temp.append("TBLU")

        print("\n database_temp: ", database_temp)
        self.model.databaseTempModel = database_temp

        for i in database_temp:
            if not(i in self.model.plc["clamps"]):
                temp = False
                break
            else:
                temp = True

        #si se colocaron las cajas necesarias de un robot se emite el ok correspondiente
        if temp:

            if self.module == "clamps_a":
                tagrob = "Robot A"
            if self.module == "clamps_b":
                tagrob = "Robot B"

            command = {
                "lbl_result" : {"text": f"Cajas de {tagrob} colocadas", "color": "green"},
                "lbl_steps" : {"text": f"Presionar boton verde o CTRL para comenzar", "color": "black"}
                }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            self.ok.emit()

class ClampsMonitorBoth(QState):
    ok = pyqtSignal()

    def __init__(self, module = "clamps", model = None, parent = None):
        super().__init__(parent)
        self.model = model
        self.module = module

    def onEntry(self, QEvent):

        temp = False
        #self.model.database["clamps"] contiene los clamps necesarios para el arnés escaneado
        #self.model.plc["clamps"] contiene las cajas que se han clampeado correctamente en físico
        print("\n database: ", self.model.database["clamps"])
        print(" PLC     : ", self.model.plc["clamps"],"\n")
        
        database_temp = []

        command = {
                    "lbl_result" : {"text": "Esperando cajas para continuar", "color": "green"},
                    "lbl_steps" : {"text": "Coloca el resto de las cajas", "color": "black"},
                    "img_center" : "logo.jpg"
                    }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        if not("PDC-D" in database_temp):
            if "PDC-D" in self.model.database["clamps"]:
                    database_temp.append("PDC-D")
        if not("PDC-P" in database_temp):
            if "PDC-P" in self.model.database["clamps"]:
                database_temp.append("PDC-P")
        if not("PDC-R" in  database_temp):
            if "PDC-R" in self.model.database["clamps"]:
                    database_temp.append("PDC-R")
        if not("PDC-RMID" in  database_temp):
            if "PDC-RMID" in self.model.database["clamps"]:
                    database_temp.append("PDC-RMID")
        if not("PDC-S" in  database_temp):
            if "PDC-S" in self.model.database["clamps"]:
                    database_temp.append("PDC-S")
        if not("PDC-S9" in  database_temp):
                if "PDC-S9" in self.model.database["clamps"]:
                    self.model.pdcs9_flag = True
                    database_temp.append("PDC-S9")
        if not("PDC-S19" in  database_temp):
            if "PDC-S19" in self.model.database["clamps"]:
                    database_temp.append("PDC-S19")
        if not("PDC-S20" in  database_temp):
            if "PDC-S20" in self.model.database["clamps"]:
                    database_temp.append("PDC-S20")
        if not("PDC-S17" in  database_temp):
            if "PDC-S17" in self.model.database["clamps"]:
                    database_temp.append("PDC-S17")
        if not("PDC-S21" in  database_temp):
            if "PDC-S21" in self.model.database["clamps"]:
                    database_temp.append("PDC-S21")
        if not("F96-1" in  database_temp):
            if "F96-1" in self.model.database["clamps"]:
                    database_temp.append("F96-1")


        if not("TBLU" in  database_temp):
            if "TBLU" in self.model.database["clamps"]:
                    database_temp.append("TBLU")

        print("\n database_temp: ", database_temp)
        self.model.databaseTempModel = database_temp

        for i in database_temp:
            if not(i in self.model.plc["clamps"]):
                temp = False
                break
            else:
                temp = True

        #si ya se colocaron todos los clamps que lleva el arnés
        if temp:

            command = {
                "lbl_result" : {"text": f"Todas las cajas colocadas", "color": "green"},
                "lbl_steps" : {"text": f"Presionar boton verde o CTRL para comenzar", "color": "black"}
                }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            self.ok.emit()

class Clamps_Standby(QState):

    def __init__(self, module = "clamps", model = None, parent = None):
        super().__init__(parent)
        self.model = model
        self.module = module

    def onEntry(self, QEvent):

        print("Esperando START")

class Finish (QState):
    ok      = pyqtSignal()
    nok     = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):

        ##################### EXCEL DE TIEMPOS #########################
        #Aquí se crea el Excel con los tiempos recopilados por los Robots
        #print("Creando excel de TIEMPOS....")
        #print("Valor Final de Tiempos: ",self.model.insertion_times)

        #wb = Workbook()
        #filestamp = strftime('%Y%m%d-%H%M%S') #Fecha y Hora actual
        #print("FILESTAMP (HORA ACTUAL): ",filestamp)
        #print("Tipo de dato de FILESTAMP: ",type(filestamp))
        #filesheet = "./"+self.model.codes["HM"]+"-"+filestamp+".xlsx" #Se crea el archivo con el nombre del HM + Fecha y Hora actual
        #sheet = wb.active
        #sheet.title = "Inserción"
        ##Se crean los encabezados
        #sheet.append(["CAVIDAD","TIEMPO_ENVIO_MENSAJE","TIEMPO_MENSAJE","TIEMPO_RESET","TIEMPO_REINICIO","TIEMPO_TOTAL","TIEMPO_ESPERA_ROBOTA","TIEMPO_ESPERA_ROBOTB","TIEMPO_TRASLADO_TOMA","TIEMPO_PRESENCIA_TOMA","TIEMPO_BAJADA_TOMA","TIEMPO_VACIO_TOMA","TIEMPO_SUBIDA_TOMA","TIEMPO_TRASLADO_INSERCION","TIEMPO_BAJADA_INSERCION","TIEMPO_CILINDRO_INSERCION","TIEMPO_INSERCION_INSERCION","TIEMPO_SUBIDA_INSERCION"])
        ##sheet["A1"] = "Cavidad"
        ##sheet["B1"] = "Tiempos"
        ##print("Data.items(): ",data.items())
        #encabezados = {}
        #print("+++++++++++Recorriendo Excel ya creado:\n")
        ##Se recorren los encabezados en el excel para identificar la columna (Ej. "A","J",etc) de cada uno y guardarlo en un arreglo (encabezados)
        #for column in range(1, sheet.max_column + 1):
        #    print("column: ",column)
        #    char = get_column_letter(column)
        #    print("Letra o Columna en Excel: ",char)
        #    header = sheet.cell(row = 1, column = column).value
        #    print("Header: ",header)
        #    encabezados[header] = char
        #print("Encabezados Arreglo: ",encabezados)
        #print("||||||self.model.insertion_times.items(): ",self.model.insertion_times.items())
        ##Se agregan al Excel las cavidades y valores correspondientes
        #for row, (key,value) in enumerate(self.model.insertion_times.items(),start=2):
        #    #print("Row dentro del for: ",row)
        #    #print("Tipo de dato del ROW: ",type(row))
        #    #print("Key dentro del for: ",key)
        #    #print("Tipo de dato KEY: ",type(key))
        #    #print("Value dentro del for: ",value)
        #    #print("Tipo de dato VALUE: ",type(value))
        #    #print("IMPRIMIENDO VALORES DENTRO DEL FOR:\n")
        #    sheet[f"A{row}"] = key
        #    for i in value:
        #        print("----i: ",i)
        #        print("value[i]",value[i])
        #        print("Columna donde se colocará: ",encabezados[i])
        #        sheet[encabezados[i]+str(row)] = str(value[i])
        #    #y = json.dumps(value)
        #    #print("VALUE CONVERTIDO (y): ",y)
        #    #print("Tipo de dato y: ",type(y))

        #wb.save(filesheet)
        #print("Excel Guardado Correctamente")

        #self.model.insertion_times = {}
        ###########################################################

        print("current state: Finish (cycle_manage)")
        command = {
                    "lbl_result" : {"text": "Guardando Información", "color": "navy"},
                    "lbl_steps" : {"text": 'Espere un momento', "color": "black"}
                    }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        Timer(0.05, self.save_info).start()


    def save_info (self):

        #se reinicia el modo de los robots para escoger en el siguiente arnés
        self.model.robots_mode = 0
        self.model.thread_robot = False
        self.model.init_thread_robot = False
        self.model.robot_principal = False
        self.model.acomodo_listas = True
        self.model.acomodo_listas_2 = True
        self.model.no_caja_actual=0
        self.model.arnes_por_finalizar=False
        self.model.cronometro_ciclo=False
        command = {}
        for i in self.model.database["fuses"]:
                command[i] = False
        publish.single(self.model.pub_topics["plc"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        historial = {
            "HM":           self.model.codes["HM"],
            "QR_FET":       self.model.codes["FET"],
            "QR_MAQUINA":   self.model.QR,
            "RESULTADO":    "BUENO",
            "FUSIBLES":     self.model.database["fuses"],
            "REINTENTOS":   self.model.retries,
            "INICIO":       self.model.datetime.isoformat(),
            "FIN":          datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
            "USUARIO":      self.model.local_data["user"]["type"] + ": " + self.model.local_data["user"]["name"] + "."
            }
        resp = requests.post(f"http://{self.model.server}/api/post/historial",data=json.dumps(historial))
        resp = resp.json()

        #### Trazabilidad FAMX2 Update de Información
        if self.model.config_data["trazabilidad"]:
            try:
                print("||Realizando el Update de SALIDA a Trazabilidad en FAMX2")
                print("ID a la que se realizará el Update para Trazabilidad",self.model.id_HM)
                salTrazabilidad = {
                    "SALINSERCION": historial["FIN"],
                    "UBICACION": "SALIDA_DE_INSERCION",
                    "NAMEINSERCION": self.model.no_serie
                    }
                endpointUpdate = "http://{}/seghm/update/seghm/{}".format(self.model.server,self.model.id_HM)
                respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(salTrazabilidad))
                respTrazabilidad = respTrazabilidad.json()
                print("respTrazabilidad del update: ",respTrazabilidad)

                if "exception" in respTrazabilidad:
                    sleep(0.5)
                    respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(salTrazabilidad))
                    respTrazabilidad = respTrazabilidad.json()
                    print("respTrazabilidad del update: ",respTrazabilidad)

                    if "exception" in respTrazabilidad:
                        sleep(0.5)
                        respTrazabilidad = requests.post(endpointUpdate, data=json.dumps(salTrazabilidad))
                        respTrazabilidad = respTrazabilidad.json()
                        print("respTrazabilidad del update: ",respTrazabilidad)

                        if "exception" in respTrazabilidad:

                            print("no se logró hacer el update en trazabilidad")
                            command = {
                                        "lbl_result" : {"text": "Problema de red: Actualización de Trazabilidad", "color": "red"},
                                        "lbl_steps" : {"text": 'Para volver a intentar botón Amarillo, para continuar "CTRL"', "color": "black"}
                                        }
                            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
                            self.model.problema_trazabilidad = True
                            print("self.model.problema_trazabilidad = True, Esperando señal de botón amarillo o tecla CTRL para reintar publish o continuar")
                            self.nok.emit()
                            return

            except Exception as ex:
                print("Excepción al momento de guardar datos en FAMX2", ex)
        #### Trazabilidad FAMX2 Update de Información


        if "HM000000011936" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True
                        
        if "HM000000011925" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True

        if "HM000000011920" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True

        if self.model.config_data["trazabilidad"] == True:
            command = {
                "lbl_info3" : {"text": "Trazabilidad\nActivada", "color": "green"}
            }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        command = {
            "lbl_result" : {"text": "Ciclo terminado", "color": "green"},
            "lbl_steps" : {"text": "Retira las cajas", "color": "black"},
            "lbl_nuts" : {"text": "", "color": "purple"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        print("self.model.modo_manual_activado finish",self.model.robothome_a)
        if self.model.modo_manual_activado==False:
            self.model.robothome_a = True # variable para activar Mensaje de enviar robot a home, se resetea sola en comm.py
            self.model.robothome_b = True # variable para activar Mensaje de enviar robot a home, se resetea sola en comm.py
        self.ok.emit()

class Reset (QState):
    ok      = pyqtSignal()
    nok     = pyqtSignal()
    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        #se reinicia variable que dice que el robot A finalizó
        self.model.robot_a_terminado = False

        self.model.cronometro_ciclo=False
        #se reinician variables por si se da llave en el momento de la inserción manual del relay
        self.model.waiting_button_inserted_singal["robot_a"] = False
        self.model.waiting_button_inserted_singal["robot_b"] = False

        if "HM000000011936" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True
                        
        if "HM000000011925" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True

        if "HM000000011920" in self.model.codes["HM"]:
            self.model.config_data["trazabilidad"] = True


        if self.model.config_data["trazabilidad"] == True:
            command = {
                "lbl_info3" : {"text": "Trazabilidad\nActivada", "color": "green"}
            }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        #se reinicia el modo de los robots para escoger en el siguiente arnés
        self.model.robots_mode = 0
        self.model.thread_robot = False
        self.model.init_thread_robot = False
        self.model.robot_principal = False
        self.model.acomodo_listas = True
        self.model.acomodo_listas_2 = True

        command = {
            "lbl_result" : {"text": "Se giró la llave de reset", "color": "green"},
            "lbl_steps" : {"text": "Reiniciando", "color": "black"},
            "lbl_nuts" : {"text": "", "color": "purple"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        command = {}
        for i in self.model.database["fuses"]:
                command[i] = False
        publish.single(self.model.pub_topics["plc"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        if self.model.datetime != None:
            historial = {
                "HM":           self.model.codes["HM"],
                "QR_FET":       self.model.codes["FET"],
                "QR_MAQUINA":   self.model.QR,
                "RESULTADO":    "MALO",
                "FUSIBLES":     self.model.database["fuses"],
                "REINTENTOS":   self.model.retries,
                "INICIO":       self.model.datetime.isoformat(),
                "FIN":          datetime.now().isoformat(),
                "USUARIO":      self.model.local_data["user"]["type"] + ": " + self.model.local_data["user"]["name"] + "."
                }
            resp = requests.post(f"http://{self.model.server}/api/post/historial",data=json.dumps(historial))
            resp = resp.json()
            if "items" in resp:
                if resp["items"] == 1:
                    pass
                else:
                    command = {
                        "lbl_result" : {"text": "Error de conexión", "color": "red"},
                        "lbl_steps" : {"text": "Datos no guardados", "color": "black"}
                        }
                    publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        Timer(2,self.ok.emit).start()

class Waiting_Robot (QState):
    ok      = pyqtSignal()
    waiting = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):

        command = {
            "lbl_info3" : {"text": "Esperando\n Segundo\n Robot", "color": "green"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

        print("Esperando segundo robot...")
        #para saber cuando el robot principal termina antes que el thread
        self.model.robot_principal = True

        #variable para saber que el robot thread ya termino
        if self.model.thread_robot == True:
            self.model.robot_principal = False
            self.ok.emit()
        elif self.model.thread_robot == False:
            Timer(2,self.waiting.emit).start()

    def onExit(self, event):
        #se limpia el mensaje de lbl_info3
        command = {
            "lbl_info3" : {"text": "", "color": "green"}
            }
        publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)

class ModoManual (QState):
    ok      = pyqtSignal()
    finish  = pyqtSignal()

    def __init__(self, model = None, parent = None):
        super().__init__(parent)
        self.model = model

    def onEntry(self, event):
        self.model.modo_manual_activado=True
        print("---------------------ModoManual--------------------------")

        #El. HMPRUEBAMANUAL01 ILX29620221004417TEST

        print("Contenido de arnés: ")
        #pprint.pprint(self.model.database["fuses"]) #también los empty
        pprint.pprint(self.model.modularity_manual)
        
        #Caja actual en la modularidad


        print("self.model.no_caja_actual",self.model.no_caja_actual)
        print(self.model.cajas_arnes)
        
        print(type(self.model.cajas_arnes))
        if self.model.no_caja_actual>=len(self.model.cajas_arnes):
            self.model.no_caja_actual=0
            self.model.arnes_por_finalizar=True


        if self.model.no_caja_actual<0:
            self.model.no_caja_actual=0
            self.model.cajas_arnes[self.model.no_caja_actual]
        caja_actual=self.model.cajas_arnes[self.model.no_caja_actual]

        

        print("caja actual: ",caja_actual)
        if self.model.arnes_recorrido==False:
            print("entro a")


            self.LlenadoDeImagen(caja_actual)
            print("Salio del llenado de la imagen")
            print("self.model.arnes_por_finalizar",self.model.arnes_por_finalizar)
            if self.model.arnes_por_finalizar==True:
                print("ENTRO AL IF")
                command = {
                "lbl_result" : {"text": "Inserta los Fusibles manualmente", "color": "green"},
                "lbl_steps" : {"text": 'Presiona "CRTL" para continuar caja o "ESC" para FINALIZAR', "color": "brown"},
                "img_center" : f"fusibles/{caja_actual}.jpg"  #Aqui actualizar la imagen principal con la siguiente CAJA
                }
                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            else:
                print("ENTRO AL ELSE")
                command = {
                    "lbl_result" : {"text": "Inserta los Fusibles manualmente", "color": "green"},
                    "lbl_steps" : {"text": 'Presiona "CRTL" para continuar caja o "flecha IZQUIERDA" para regresar', "color": "black"},
                    "img_center" : f"fusibles/{caja_actual}.jpg"  #Aqui actualizar la imagen principal con la siguiente CAJA
                    }
                publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
        if self.model.confirmacion_arnes_finalizado==True and self.model.confirmacion_arnes_finalizado==True:
            command = {
                "lbl_result" : {"text": "Finalizando Inserción", "color": "green"},
                "lbl_steps" : {"text": 'Guardando Información', "color": "navy"},
                "img_center" : "logo.jpg"
                }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            print("Finalizar", self.model.arnes_recorrido)
            self.model.confirmacion_arnes_finalizado=False
            self.model.arnes_finalizado=False
            Timer(1,self.finish.emit).start()
            #DESCOMENTAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAR PARA FINALIZAR MODO MANUAL
        
    def LlenadoDeImagen(self,caja_actual):
        

        fusibles_verticales = {"F400","F401","F402","F403","F404","F405","F406","F407","F408","F409","F410","F411",
                               "F430","F431","F432","F433","F436",
                               "F437","F438","F439","F440","F441","F442","F443","F444","F445","F446",
                               "F450","F451","F452","F453","F454","F455","F456","F457","F458","F459","F460","F461",
                               "1","2","3","4","5","6",
                               "1","2","3","4","5","6","7","8","9"}

        #leer imagen de caja
        imgcaja = cv2.imread(self.model.imgs_path + "/fusibles/cajas/" + str(caja_actual) + ".jpg")
        ancho_caja = imgcaja.shape[1] #columnas
        alto_caja = imgcaja.shape[0] #filas

        for elemento in self.model.modularity_manual_respaldo:
            if caja_actual in elemento[0]:
                print(elemento)
                #leer imagen de fusible
                imgfusible = cv2.imread(self.model.imgs_path + "/fusibles/" + str(elemento[2]) + ".jpg")

                if elemento[1] in fusibles_verticales:
                    #Rotar imagen
                    # Using cv2.rotate() method
                    # Using cv2.ROTATE_90_CLOCKWISE rotate by 90 degrees clockwise
                    # Using cv2.ROTATE_180 rotate by 180 degrees clockwise
                    # Using cv2.ROTATE_90_COUNTERCLOCKWISE rotate by 270 degrees clockwise
                    imgfusible = cv2.rotate(imgfusible, cv2.cv2.ROTATE_90_COUNTERCLOCKWISE)

                X=100
                Y=100

                if "PDC-R" in elemento[0]:
                    if "F400" in elemento[1]:
                        X=2870
                        Y=1060
                    if "F401" in elemento[1]:
                        X=2786
                        Y=1060
                    if "F402" in elemento[1]:
                        X=2702
                        Y=1060
                    if "F403" in elemento[1]:
                        X=2618
                        Y=1060
                    if "F404" in elemento[1]:
                        X=2534
                        Y=1060
                    if "F405" in elemento[1]:
                        X=2440
                        Y=1060


                    if "F406" in elemento[1]:
                        X=2260
                        Y=1100
                    if "F407" in elemento[1]:
                        X=2176
                        Y=1100
                    if "F408" in elemento[1]:
                        X=2092
                        Y=1100
                    if "F409" in elemento[1]:
                        X=2008
                        Y=1100
                    if "F410" in elemento[1]:
                        X=1924
                        Y=1100
                    if "F411" in elemento[1]:
                        X=1830
                        Y=1100

                    if "F450" in elemento[1]:
                        X=2870
                        Y=220
                    if "F451" in elemento[1]:
                        X=2784
                        Y=220
                    if "F452" in elemento[1]:
                        X=2698
                        Y=220
                    if "F453" in elemento[1]:
                        X=2612
                        Y=220
                    if "F454" in elemento[1]:
                        X=2526
                        Y=220
                    if "F455" in elemento[1]:
                        X=2440
                        Y=220


                    if "F456" in elemento[1]:
                        X=2260
                        Y=220
                    if "F457" in elemento[1]:
                        X=2176
                        Y=220
                    if "F458" in elemento[1]:
                        X=2092
                        Y=220
                    if "F459" in elemento[1]:
                        X=2008
                        Y=220
                    if "F460" in elemento[1]:
                        X=1924
                        Y=220
                    if "F461" in elemento[1]:
                        X=1830
                        Y=220


                    if "F412" in elemento[1]:
                        X=2990
                        Y=1120
                    if "F413" in elemento[1]:
                        X=2990
                        Y=1057
                    if "F414" in elemento[1]:
                        X=2990
                        Y=994
                    if "F415" in elemento[1]:
                        X=2990
                        Y=931
                    if "F416" in elemento[1]:
                        X=2990
                        Y=868
                    if "F417" in elemento[1]:
                        X=2990
                        Y=805


                    if "F421" in elemento[1]:
                        X=2990
                        Y=665
                    if "F422" in elemento[1]:
                        X=2990
                        Y=602
                    if "F423" in elemento[1]:
                        X=2990
                        Y=539
                    if "F424" in elemento[1]:
                        X=2990
                        Y=476
                    if "F425" in elemento[1]:
                        X=2990
                        Y=413
                    if "F426" in elemento[1]:
                        X=2990
                        Y=350


                    if "F418" in elemento[1]:
                        X=1390
                        Y=1100
                    if "F419" in elemento[1]:
                        X=1390
                        Y=950
                    if "F420" in elemento[1]:
                        X=1390
                        Y=810


                    if "F447" in elemento[1]:
                        X=1390
                        Y=620
                    if "F448" in elemento[1]:
                        X=1390
                        Y=480
                    if "F449" in elemento[1]:
                        X=1390
                        Y=340


                    if "RELT" in elemento[1]:
                        X=1810
                        Y=750
                    if "RELU" in elemento[1]:
                        X=2185
                        Y=750
                    if "RELX" in elemento[1]:
                        X=2595
                        Y=750

                    if "F430" in elemento[1]:
                        X=2495
                        Y=600
                    if "F431" in elemento[1]:
                        X=2410
                        Y=600
                    if "F432" in elemento[1]:
                        X=2220
                        Y=600
                    if "F433" in elemento[1]:
                        X=2135
                        Y=600
                    if "F436" in elemento[1]:
                        X=1880
                        Y=600

                    if "F437" in elemento[1]:
                        X=2750
                        Y=455
                    if "F438" in elemento[1]:
                        X=2665
                        Y=455
                    if "F439" in elemento[1]:
                        X=2580
                        Y=455
                    if "F440" in elemento[1]:
                        X=2495
                        Y=455
                    if "F441" in elemento[1]:
                        X=2410
                        Y=455

                    if "F442" in elemento[1]:
                        X=2220
                        Y=455
                    if "F443" in elemento[1]:
                        X=2135
                        Y=455
                    if "F444" in elemento[1]:
                        X=2050
                        Y=455
                    if "F445" in elemento[1]:
                        X=1965
                        Y=455
                    if "F446" in elemento[1]:
                        X=1880
                        Y=455


                if "F96" in elemento[0]:
                    if "F96" in elemento[1]:
                        X=192
                        Y=238

                if "PDC-S" in elemento[0]:
                    if "1" in elemento[1]:
                        X=437
                        Y=195
                    if "2" in elemento[1]:
                        X=387
                        Y=195
                    if "3" in elemento[1]:
                        X=337
                        Y=195
                    if "4" in elemento[1]:
                        X=287
                        Y=195
                    if "5" in elemento[1]:
                        X=237
                        Y=195
                    if "6" in elemento[1]:
                        X=187
                        Y=195

                if "TBLU" in elemento[0]:
                    if "1" in elemento[1]:
                        X=1163
                        Y=240
                    if "2" in elemento[1]:
                        X=1062
                        Y=240
                    if "3" in elemento[1]:
                        X=961
                        Y=240
                    if "4" in elemento[1]:
                        X=860
                        Y=240
                    if "5" in elemento[1]:
                        X=759
                        Y=240
                    if "6" in elemento[1]:
                        X=658
                        Y=240
                    if "7" in elemento[1]:
                        X=557
                        Y=240
                    if "8" in elemento[1]:
                        X=456
                        Y=240
                    if "9" in elemento[1]:
                        X=355
                        Y=240

                if "PDC-D" in elemento[0]:
                    
                    if "F209" in elemento[1]:
                        X=1150
                        Y=1740
                    if "F210" in elemento[1]:
                        X=1150
                        Y=1640
                    if "F211" in elemento[1]:
                        X=1150
                        Y=1540
                    if "F212" in elemento[1]:
                        X=1150
                        Y=1440
                    if "F213" in elemento[1]:
                        X=1150
                        Y=1340
                    if "F214" in elemento[1]:
                        X=1150
                        Y=1240
                    if "F215" in elemento[1]:
                        X=1150
                        Y=1140
                    if "F216" in elemento[1]:
                        X=1150
                        Y=1040


                    if "F208" in elemento[1]:
                        X=463
                        Y=950
                    if "F207" in elemento[1]:
                        X=463
                        Y=1050
                    if "F206" in elemento[1]:
                        X=463
                        Y=1150
                    if "F205" in elemento[1]:
                        X=463
                        Y=1250
                    if "F204" in elemento[1]:
                        X=463
                        Y=1350
                    if "F203" in elemento[1]:
                        X=463
                        Y=1450
                    if "F202" in elemento[1]:
                        X=463
                        Y=1550
                    if "F201" in elemento[1]:
                        X=463
                        Y=1650
                    if "F200" in elemento[1]:
                        X=463
                        Y=1750

                    if "F221" in elemento[1]:
                        X=617
                        Y=330
                    if "F220" in elemento[1]:
                        X=617
                        Y=430
                    if "F219" in elemento[1]:
                        X=617
                        Y=530
                    if "F218" in elemento[1]:
                        X=617
                        Y=630
                    if "F217" in elemento[1]:
                        X=617
                        Y=730
                    
                    if "F226" in elemento[1]:
                        X=858
                        Y=330
                    if "F225" in elemento[1]:
                        X=858
                        Y=430
                    if "F224" in elemento[1]:
                        X=858
                        Y=530
                    if "F223" in elemento[1]:
                        X=858
                        Y=630
                    if "F222" in elemento[1]:
                        X=858
                        Y=730

                    if "F232" in elemento[1]:
                        X=1200
                        Y=330
                    if "F231" in elemento[1]:
                        X=1200
                        Y=430
                    if "F230" in elemento[1]:
                        X=1200
                        Y=530
                    if "F229" in elemento[1]:
                        X=1200
                        Y=610
                    if "F228" in elemento[1]:
                        X=1200
                        Y=710
                    if "F227" in elemento[1]:
                        X=1200
                        Y=810


                if "PDC-P" in elemento[0]:
                    
                    if "MF1" in elemento[1]:
                        X=800
                        Y=350
                    if "MF2" in elemento[1]:
                        X=800
                        Y=490


                    if "F326" in elemento[1]:
                        X=1346
                        Y=1320
                    if "F327" in elemento[1]:
                        X=1346
                        Y=1227
                    if "F328" in elemento[1]:
                        X=1346
                        Y=1134
                    if "F329" in elemento[1]:
                        X=1346
                        Y=1041
                    if "F330" in elemento[1]:
                        X=1346
                        Y=948
                    if "F331" in elemento[1]:
                        X=1346
                        Y=855
                    if "F332" in elemento[1]:
                        X=1346
                        Y=762
                    if "F333" in elemento[1]:
                        X=1346
                        Y=669
                    if "F334" in elemento[1]:
                        X=1346
                        Y=576
                    if "F335" in elemento[1]:
                        X=1340
                        Y=480

                    if "F318" in elemento[1]:
                        X=1138
                        Y=1320
                    if "F319" in elemento[1]:
                        X=1138
                        Y=1226
                    if "F320" in elemento[1]:
                        X=1138
                        Y=1132
                    if "F321" in elemento[1]:
                        X=1138
                        Y=1038
                    if "F322" in elemento[1]:
                        X=1138
                        Y=944
                    if "F323" in elemento[1]:
                        X=1138
                        Y=850
                    if "F324" in elemento[1]:
                        X=1138
                        Y=756
                    if "F325" in elemento[1]:
                        X=1138
                        Y=665

                    if "F300" in elemento[1]:
                        X=800
                        Y=1100

                    if "F301" in elemento[1]:
                        X=840
                        Y=1025
                    if "F302" in elemento[1]:
                        X=840
                        Y=925
                    if "F303" in elemento[1]:
                        X=840
                        Y=826
                    if "F304" in elemento[1]:
                        X=840
                        Y=732
                    if "F305" in elemento[1]:
                        X=840
                        Y=635

                #PDCRMID:
                if "PDC-RMID" in elemento[0]:

                    if "F400" in elemento[1]:
                        X=2896
                        Y=1140
                    if "F401" in elemento[1]:
                        X=2780
                        Y=1140
                    if "F402" in elemento[1]:
                        X=2669
                        Y=1140
                    if "F403" in elemento[1]:
                        X=2558
                        Y=1140
                    if "F404" in elemento[1]:
                        X=2437
                        Y=1140
                    if "F405" in elemento[1]:
                        X=2320
                        Y=1140


                    if "F406" in elemento[1]:
                        X=2067
                        Y=1190
                    if "F407" in elemento[1]:
                        X=1948
                        Y=1190
                    if "F408" in elemento[1]:
                        X=1830
                        Y=1190
                    if "F409" in elemento[1]:
                        X=1712
                        Y=1190
                    if "F410" in elemento[1]:
                        X=1594
                        Y=1190
                    if "F411" in elemento[1]:
                        X=1475
                        Y=1190

                    if "F450" in elemento[1]:
                        X=2896
                        Y=-50
                    if "F451" in elemento[1]:
                        X=2780
                        Y=-50
                    if "F452" in elemento[1]:
                        X=2669
                        Y=-50
                    if "F453" in elemento[1]:
                        X=2558
                        Y=-50
                    if "F454" in elemento[1]:
                        X=2437
                        Y=-50
                    if "F455" in elemento[1]:
                        X=2320
                        Y=-50


                    if "F456" in elemento[1]:
                        X=2067
                        Y=-50
                    if "F457" in elemento[1]:
                        X=1948
                        Y=-50
                    if "F458" in elemento[1]:
                        X=1830
                        Y=-50
                    if "F459" in elemento[1]:
                        X=1712
                        Y=-50
                    if "F460" in elemento[1]:
                        X=1594
                        Y=-50
                    if "F461" in elemento[1]:
                        X=1475
                        Y=-50


                    if "F412" in elemento[1]:
                        X=3080
                        Y=1220
                    if "F413" in elemento[1]:
                        X=3080
                        Y=1135
                    if "F414" in elemento[1]:
                        X=3080
                        Y=1050
                    if "F415" in elemento[1]:
                        X=3080
                        Y=965
                    if "F416" in elemento[1]:
                        X=3080
                        Y=880
                    if "F417" in elemento[1]:
                        X=3080
                        Y=795


                    if "F421" in elemento[1]:
                        X=3080
                        Y=590
                    if "F422" in elemento[1]:
                        X=3080
                        Y=505
                    if "F423" in elemento[1]:
                        X=3080
                        Y=420
                    if "F424" in elemento[1]:
                        X=3080
                        Y=335
                    if "F425" in elemento[1]:
                        X=3080
                        Y=250
                    if "F426" in elemento[1]:
                        X=3080
                        Y=165


                    if "F418" in elemento[1]:
                        X=865
                        Y=1190
                    if "F419" in elemento[1]:
                        X=865
                        Y=990
                    if "F420" in elemento[1]:
                        X=865
                        Y=790


                    if "F447" in elemento[1]:
                        X=865
                        Y=520
                    if "F448" in elemento[1]:
                        X=865
                        Y=320
                    if "F449" in elemento[1]:
                        X=865
                        Y=120


                    if "RELT" in elemento[1]:
                        X=1460
                        Y=705
                    if "RELU" in elemento[1]:
                        X=1970
                        Y=705
                    if "RELX" in elemento[1]:
                        X=2533
                        Y=705

                    if "F430" in elemento[1]:
                        X=2402
                        Y=470
                    if "F431" in elemento[1]:
                        X=2286
                        Y=470
                    if "F432" in elemento[1]:
                        X=2025
                        Y=470
                    if "F433" in elemento[1]:
                        X=1909
                        Y=470
                    if "F436" in elemento[1]:
                        X=1556
                        Y=470

                    if "F437" in elemento[1]:
                        X=2750
                        Y=290
                    if "F438" in elemento[1]:
                        X=2634
                        Y=290
                    if "F439" in elemento[1]:
                        X=2518
                        Y=290
                    if "F440" in elemento[1]:
                        X=2402
                        Y=290
                    if "F441" in elemento[1]:
                        X=2286
                        Y=290

                    if "F442" in elemento[1]:
                        X=2025
                        Y=290
                    if "F443" in elemento[1]:
                        X=1904
                        Y=290
                    if "F444" in elemento[1]:
                        X=1788
                        Y=290
                    if "F445" in elemento[1]:
                        X=1672
                        Y=290
                    if "F446" in elemento[1]:
                        X=1556
                        Y=290

                    movx = -30
                    movy = 360

                    X = X + movx
                    Y = Y + movy

                #print("X: ",X)
                #print("Y: ",Y)
                #obtener ancho y alto de imágen de fusible
                ancho_fusible = imgfusible.shape[1] #columnas
                alto_fusible = imgfusible.shape[0] #filas
                if "PDC-RMID" in elemento[0]  or  "PDC-D" in elemento[0] or  "PDC-P" in elemento[0]:
                    scale_percent = 170 # percent of original size
                else:
                    scale_percent = 100 # percent of original size
                if elemento[1] in fusibles_verticales:
                    width = int(imgfusible.shape[1] * scale_percent / 80)
                    height = int(imgfusible.shape[0] * scale_percent / 100)
                else:
                    width = int(imgfusible.shape[1] * scale_percent / 80)
                    height = int(imgfusible.shape[0] * scale_percent / 90)
                if "RELT" in elemento[1] or "RELU" in elemento[1] or "RELX" in elemento[1]:
                    width = int(imgfusible.shape[1] * scale_percent / 80)
                    height = int(imgfusible.shape[0] * scale_percent / 100)
                
                if "PDC-P" in  elemento[0]:
                    width = int(imgfusible.shape[1] * scale_percent / 100)
                    height = int(imgfusible.shape[0] * scale_percent / 90)
                
                if "F96" in elemento[0] or  "PDC-S" in elemento[0]:
                    width = int(imgfusible.shape[1] * scale_percent / 100)
                    height = int(imgfusible.shape[0] * scale_percent /100)

                if "PDC-R" in elemento[0] and not "PDC-RMID" in elemento[0]:
                    width = int(imgfusible.shape[1] * scale_percent / 65)
                    height = int(imgfusible.shape[0] * scale_percent / 80)

                dim = (width, height)
                  
                # resize image
                imgfusible = cv2.resize(imgfusible, dim, interpolation = cv2.INTER_AREA)
                w=width
                h=height
                #imgfusible = cv2.resize(imgfusible,(w,h))
                imgcaja[Y:Y+h,X:X+w] = imgfusible

                
        
        
        #Guardar imagen
        cv2.imwrite(self.model.imgs_path + f"/fusibles/{caja_actual}.jpg", imgcaja)

    def onExit(self, event):
        print("saliendo de ModoManual")

        if self.model.arnes_recorrido==False:
            command = {
                "lbl_result" : {"text": "Caja Finalizada", "color": "green"},
                "lbl_steps" : {"text": 'Avanzando a Siguiente Caja', "color": "black"},
                }
            publish.single(self.model.pub_topics["gui"],json.dumps(command),hostname='127.0.0.1', qos = 2)
            
            caja_actual =self.model.cajas_arnes[self.model.no_caja_actual]
            if self.model.regreso==True:
                self.model.no_caja_actual-=1
                self.model.regreso=False
            if self.model.control_presionado==True:
                self.model.no_caja_actual+=1
                self.model.control_presionado=False
            
            
            if not caja_actual in self.model.cajas:
                self.model.cajas.append(caja_actual)
            print("self.model.cajas",self.model.cajas)
            
 
        