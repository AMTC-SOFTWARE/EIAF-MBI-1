
"""
@author: MS. Marco Rutiaga Quezada
         MS. Aarón Castillo Tobías

Upload file. Basic front code:
    <!doctype html>
    <title>Upload new File</title>
    <h1>Upload new File</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=file>
      <input type=submit value=Upload>
    </form>
    ###############################################################################
        command to exe generation_
        pyinstaller --noconfirm api.py
        pyinstaller --noconsole --icon=icon.ico --add-data data;data api.py
        pyinstaller --icon=icon.ico --add-data data;data api.py
        python -m PyInstaller --icon=icon.ico --add-data data;data api.py
        pyinstaller --onedir --icon=icon.ico --contents-directory "." --add-data 'data;data' api.py
###############################################################################
"""

import os
from flask import Flask,  request,send_file, make_response
from werkzeug.utils import secure_filename
from flask_cors import CORS
import requests
import io
from openpyxl import Workbook
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date, time
from time import strftime
import pymysql
from os.path import exists  #para saber si existe una carpeta o archivo
from shutil import rmtree   #para eliminar carpeta con archivos dentro: rmtree("carpeta_con_archivos")
from os import remove       #para eliminar archivo único: remove("archivo.txt")
from os import rmdir        #para eliminar carpeta vacía: rmdir("carpeta_vacia")
import json
import pyodbc # Librería que permite conexión con FAMX2
import auto_modularities
from model import model


datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2=datos_conexion.datos_acceso()

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), '..\\')

@app.route("/server_famx/hora_servidor",methods=["GET"])
def servidorHora():
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("Conexión Éxitosa")
    except Exception as ex:
        print("Conexión a P2 Exception: ", ex)
        return {"exception": ex.args}
    query = "SELECT CURRENT_TIMESTAMP AS HORA_ACTUAL;"
    print("query: ",query)
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            #result = cursor.fetchall()
            records = cursor.fetchall()
            insertObject = []
            columnNames = [column[0]
               for column in cursor.description
            ]
            for record in records:
               insertObject.append(dict(zip(columnNames, record)))
               #print("insertObject FINAL: ",insertObject)
            if len(insertObject) == 1:
                response = insertObject[0]
            elif len(insertObject) > 1:
                response = {}
                keys = list(insertObject[0])
                for key in keys:
                    response[key] = []
                    for item in insertObject:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": 0}
            if "HORA_ACTUAL" in response:
                response["HORA_ACTUAL"] = response["HORA_ACTUAL"].strftime('%Y-%m-%d %H:%M:%S')
                print("response[HORA_ACTUAL]: ",response["HORA_ACTUAL"])
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response
#####################################  Upload Files Services ####################################
@app.route('/delete/filesmodularities', methods=['POST'])
def delRef():
    response = {"items": 0}
    try:
        path_carpeta = "..\\ILX";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                #rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
                response = {"path" : 'Carpeta Eliminada desde la API,'}
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)
                response = {"exception" : ex.args}
    except Exception as ex:
        print("uploadRef Exception: ", ex)
        response = {"exception" : ex.args}
        return response


@app.route('/upload/modularities', methods=['POST'])
def uploadRef():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() == "dat"
        if file and allowed_file:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], 'ILX')
            print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')

            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "ILX", filename))
            response["items"] = 1
    except Exception as ex:
        print("uploadRef Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

@app.route('/update/modularities', methods=['POST'])
def updateRef():
    data = request.form['DBEVENT']
    print("DB a la que se cargn los DAT: ",data)
    ilxfaltantes = auto_modularities.makeModularities(data)
    return ilxfaltantes

@app.route('/update/modules', methods=['POST'])
def updateModules():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        path_carpeta = "..\\modules";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)

        data = request.form['DBEVENT']
        print("DB a la que se carga la Info: ",data)
        usuario = request.form['USUARIO']
        print("Usuario que carga la info: ",usuario)
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']
        if file and allowed_file:
            filename = secure_filename(file.filename)

            path = os.path.join(app.config['UPLOAD_FOLDER'], "modules")
            #print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')
            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "modules", filename))
            auto_modularities.refreshModules(data)
            excelnew = {
                'DBEVENT': data,
                'ARCHIVO': filename,
                'USUARIO': usuario,
                'DATETIME': 'AUTO'
                }
            print("Información que se manda al POST DE EVENTOS HISTORIAL: ",excelnew)
            endpoint = f"http://{host}:5000/api/post/historial"
            responseHistorial = requests.post(endpoint, data = json.dumps(excelnew))
            response["items"] = 1
    except Exception as ex:
        print("updateModules Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

@app.route('/update/determinantes', methods=['POST'])
def updateDeterminantes():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        path_carpeta = "..\\determinantes";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)

        data = request.form['DBEVENT']
        print("DB a la que se carga la Info: ",data)
        usuario = request.form['USUARIO']
        print("Usuario que carga la info: ",usuario)
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']
        if file and allowed_file:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], "determinantes")
            print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')
            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "determinantes", filename))
            auto_modularities.refreshDeterminantes(data,usuario)
            response["items"] = 1
    except Exception as ex:
        print("updateDeterminantes Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

#########################################  CRUD Services ########################################
@app.route("/api/get/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def GET(table, column_1, operation_1, value_1, column_2, operation_2, value_2):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("GET connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            if len(result) > 0:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    for item in result:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": items}
    except Exception as ex:
        print("GET cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/post/<table>",methods=["POST"])
def POST(table):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    try:
        if ("DBEVENT" in data):
            #print("True SI HAY DBEVENT")
            print("DBEVENT: ",data["DBEVENT"])
            connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
            del data["DBEVENT"]
        else:
            #print ("False NO HAY DBEVENT, TODO FLUYE NORMAL")
            connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("POST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "INSERT INTO " + table
        keys = list(data)
        cols = ', '.join(map(escape_name, keys))
        placeholders = ', '.join(['%({})s'.format(key) for key in keys])
        query += ' ({}) VALUES ({})'.format(cols, placeholders)
        for key in data:
            try:
                if key == "FECHA" or key == "DATETIME":
                    if data[key] == "AUTO":
                        data[key] = datetime.now().isoformat()
                if type(data[key]) == dict:
                    data[key] = json.dumps(data[key])
            except Exception as ex:
                print("keys inspection Exception: ", ex)
        with connection.cursor() as cursor:
            items = cursor.execute(query, data)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("POST Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/delete/<table>/<int:ID>",methods=["POST"])
def DELETE(table, ID):
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("DELETE connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(f"DELETE FROM {table} WHERE ID={ID}")
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("DELETE Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/update/<table>/<int:ID>",methods=["POST"])
def UPDATE(table, ID):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    try:
        if ("DBEVENT" in data):
            #print("True SI HAY DBEVENT")
            print("DBEVENT: ",data["DBEVENT"])
            connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
            del data["DBEVENT"]
        else:
            #print ("False NO HAY DBEVENT, TODO FLUYE NORMAL")
            connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("UPDATE connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "UPDATE " + table + f" SET"
        for i in data:
            if i == "FECHA" or i == "DATETIME":
                if data[i] == "AUTO":
                    data[i] = datetime.now().isoformat()
            key = escape_name(i)
            if type(data[i]) == dict:
                data[i] = json.dumps(data[i])
            query += f' {key}=%({i})s,'
        query = query[:-1]
        query += f" WHERE ID={ID}"
        with connection.cursor() as cursor:
            items = cursor.execute(query,data)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("UPDATE Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response


@app.route("/api/get/pdcr/variantes",methods=["GET"])
def variantes():
    pdcrVariantes = {
    "small": [],
    "medium": [],
    "large": [],
    "battery-2": []
    }

    endpoint = f"http://{host}:5000/api/get/definiciones/ACTIVO/=/1/_/_/_"
    pdcrVariantesDB = requests.get(endpoint).json()
    if "exception" in pdcrVariantesDB:
        endpoint = f"http://{host}:5000/api/get/definiciones/ACTIVE/=/1/_/_/_"
        pdcrVariantesDB = requests.get(endpoint).json()

    #print("pdcrVariantesDB-------",pdcrVariantesDB)
    if len(pdcrVariantesDB["MODULO"]) > 0:
        #print("Cantidad de Módulos: ",len(pdcrVariantesDB["MODULO"]))
        #print("Lista de Módulos: ",pdcrVariantesDB["MODULO"])
        #print("Lista de Variantes: ",pdcrVariantesDB["VARIANTE"])
        for i in pdcrVariantesDB["MODULO"]:
            #print("Modulo Actual (i)",i)
            #print("Index de Modulo Actual (i)",pdcrVariantesDB["MODULO"].index(i))
            #print("Variante correspondiente a Modulo Actual: ",pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)])
            if pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-R":
                pdcrVariantes["large"].append(i)
                #print("ES UNA PDC-R LARGE")
            elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RMID":
                #print("ES UNA PDC-R MEDIUM")
                pdcrVariantes["medium"].append(i)
            elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RS":
                #print("ES UNA PDC-R SMALL")
                pdcrVariantes["small"].append(i)
            elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "BATTERY-2":
                pdcrVariantes["battery-2"].append(i)
    return pdcrVariantes

################################################## Respaldos de Base de Datos Endpoint  ####################################################
@app.route("/api/get/bkup",methods=["GET"])
def bkup():
    items = {
        "status": False,
        "dir": "",
        "nombre": ""
        }
    ####### Cambiar Dirección de la carpeta destino donde se guardarán los Backups, dependiendo de la máquina o computadora en la que se correrá la API #######
    dest_folder = "C:/Users/EIAF-MBI/Documents/EIAF_BKUPS/DATABASE"
    print("Petición de BACKUP")
    try:
        if os.path.isdir(dest_folder):
            print("La Carpeta para respaldos SI existe!")
            path = os.getcwd()   # show current working directory (cwd)
            print("path",path)
            os.chdir('C:/xampp/mysql/bin')
            filestamp = strftime('%Y%m%d-%H%M%S')
            filename = "%s/%s-%s.sql" % (dest_folder, filestamp, database)
            db_dump = "mysqldump --single-transaction -h " + host + " -u " + user + " -p" + password + " " + database + " > " + filename
            os.system(db_dump)
            items["status"] = True
            items["dir"] = filename
            items["nombre"] = filestamp+"-"+database
            print("DATABASE BACKUP EXITOSO")
        else:
            print("La Carpeta para respaldos NO existe!")
            items["dir"] = dest_folder
    except Exception as ex:
        print("DB BKUP Exception: ",ex)
    return items

################################################## Crear Base de Datos (Evento)  ####################################################
@app.route("/api/post/newEvent",methods=["POST"])
def newEvent():
    host_fase = host
    user_fase = user
    password_fase = "4dm1n_001"
    charSet = "utf8mb4_bin"
    historial = {
        "DBEVENT": "",
        "ARCHIVO": "",
        "USUARIO": "",
        "DATETIME": "",
    }
    activo = {
        "DBEVENT": "",
        "ACTIVO": ""
    }

    data = request.get_json(force=True)
    print("Data: ",data)
    event_name = 'evento_'+data["EVENTO"]+"_"+data["NUMERO"]+"_"+data["CONDUCCION"]
    historial["USUARIO"] = data["USUARIO"]
    historial["DATETIME"] = data["DATETIME"]
    historial["DBEVENT"] = event_name
    print(data)
    escaped_event_name = f"`{event_name}`"

    if "ACTIVO" in data:
        activo["ACTIVO"] = data["ACTIVO"]
    elif "ACTIVE" in data:
        activo["ACTIVO"] = data["ACTIVE"]

    activo["DBEVENT"] = event_name
    try:
        connection = pymysql.connect(host = host_fase, user = user_fase, passwd = password_fase)
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("create database "+escaped_event_name)
            sql = "use "+escaped_event_name
            cursor.execute(sql)
            definicionesTable = """CREATE TABLE definiciones (
            ID int primary key AUTO_INCREMENT,
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            VARIANTE text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            DATETIME datetime NOT NULL,
            USUARIO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            ACTIVO tinyint NOT NULL
            )"""
            cursor.execute(definicionesTable)
            fusiblesTable = """CREATE TABLE modulos_fusibles (
            ID int primary key AUTO_INCREMENT,
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-R` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-RMID` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-RS` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S9` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S19` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S20` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S21` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-S17` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `F96` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `F96-1` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `TBLU` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-D` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            `PDC-P` longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL
            )"""
            cursor.execute(fusiblesTable)
            modularidadesTable = """CREATE TABLE modularidades (
            ID int primary key AUTO_INCREMENT, 
            MODULARIDAD text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            FECHA datetime NOT NULL,
            MODULOS_FUSIBLES text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            ACTIVO tinyint NOT NULL
            )"""
            cursor.execute(modularidadesTable)
            historialTable = """CREATE TABLE historial (
            ID int primary key AUTO_INCREMENT, 
            ARCHIVO longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            USUARIO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            DATETIME datetime NOT NULL
            )"""
            cursor.execute(historialTable)
            activoTable = """CREATE TABLE activo (
            ID int primary key AUTO_INCREMENT, 
            ACTIVO tinyint NOT NULL
            )"""
            cursor.execute(activoTable)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("generalPOST Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        #print("Información que se manda al POST DE EVENTOS HISTORIAL: ",historial)
        endpoint = f"http://{host}:5000/api/post/historial"
        responseHistorial = requests.post(endpoint, data = json.dumps(historial))
        #print("Información que se manda al POST DE EVENTOS ACTIVO: ",activo)
        endpoint = f"http://{host}:5000/api/post/activo"
        responseActivo = requests.post(endpoint, data = json.dumps(activo))
        connection.close()
        return response

################################################## Eliminar Base de Datos (Evento)  ####################################################
@app.route("/api/delete/event",methods=["POST"])
def delEvent():
    charSet = "utf8mb4_bin"
    response = {"delete": 0}

    data = request.get_json(force=True)
    print("Data: ",data)
    #EVENTDELETE = data["DBEVENT"]
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
    except Exception as ex:
        print("Delete Event connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("DROP DATABASE "+data["DBEVENT"])
        connection.commit()
        response["delete"] = 1
    except Exception as ex:
        print("Delete Event Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response
################################################## Consultar Bases de Datos (Eventos)  ####################################################
@app.route("/api/get/eventos",methods=["GET"])
def eventos():
    lista = {
        "eventos": {}
        }
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password)
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("SHOW DATABASES")
            l = cursor.fetchall()
            #print ("Lista de dbs: ",l)
            x = []
            for i in l:
                #print("imprimiendo I 0 ",i[0])
                if 'evento' in i[0]:
                    #print("Este contiene evento: ",i[0])
                    x.extend(i)
                    
                    endpoint = f"http://{host}:5000/api/get/{i[0]}/historial/all/-/-/-/-/-"
                    respHistorial = requests.get(endpoint).json()
                    endpoint = f"http://{host}:5000/api/get/{i[0]}/activo/all/-/-/-/-/-"
                    respActivo = requests.get(endpoint).json()
                    if "exception" in respActivo:
                        endpoint = f"http://{host}:5000/api/get/{i[0]}/active/all/-/-/-/-/-"
                        respActivo = requests.get(endpoint).json()

                    #print("Respuesta de Historial: ",respHistorial)
                    #print("Respuesta de Historial Archivo: ",respHistorial["ARCHIVO"])
                    #print("Respuesta de Activo: ",respActivo)
                    #print("Respuesta de Activo: ",respActivo["ACTIVO"])

                    if "ACTIVO" in respActivo:
                        respuestaActivoo = respActivo["ACTIVO"]
                    elif "ACTIVE" in respActivo:
                        respuestaActivoo = respActivo["ACTIVE"]

                    if type(respHistorial["ARCHIVO"]) == list:
                        #print("Es una lista!")
                        lista["eventos"][i[0]] = [respHistorial["ARCHIVO"][-1],respuestaActivoo]
                    else:
                        #print("No es una lista, es posible que sea solo un elemento o esté vacío")
                        lista["eventos"][i[0]] = [respHistorial["ARCHIVO"],respuestaActivoo]
            #print("Lista de bases de datos: ",x)
            print("Lista de eventos final: ",lista)
        connection.commit()
    except Exception as ex:
        print("generalPOST Exception: ", ex)
    finally:
        connection.close()
        return lista

@app.route("/api/get/<db>/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def eventGET(table, column_1, operation_1, value_1, column_2, operation_2, value_2, db):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = db, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("GET connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            if len(result) > 0:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    for item in result:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": items}
    except Exception as ex:
        print("GET cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/get/<db>/preview/modularity/<ILX>",methods=["GET"])
def previewEvent(ILX,db):
     endpoint = f"http://{host}:5000/api/get/{db}/pdcr/variantes"
     pdcrVariantes = requests.get(endpoint).json()
     print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
     flag_l = False
     flag_m = False
     flag_s = False
     endpoint = f"http://{host}:5000/api/get/{db}/modularidades/MODULARIDAD/=/{ILX}/ACTIVO/=/1"
     response = requests.get(endpoint).json()
     if "exception" in response:
         endpoint = f"http://{host}:5000/api/get/{db}/modularidades/MODULARIDAD/=/{ILX}/ACTIVE/=/1"
         response = requests.get(endpoint).json()
     #arrayModules = response["MODULOS_FUSIBLES"][0].split(",")
     modules = response["MODULOS_FUSIBLES"][0].split(sep = ",")
     print(f"\n\t\tMODULOS_FUSIBLES:\n{modules}")
     #print("Modulos SPLIT: ",arrayModules)
     modularity = {
         'PDC-P': {},
         'PDC-D': {},
         'PDC-R': {},
         'PDC-RMID': {},
         'PDC-RS': {},
         'PDC-S': {}, 
         'PDC-S9': {}, 
         'PDC-S19': {}, 
         'PDC-S20': {}, 
         'PDC-S21': {}, 
         'PDC-S17': {}, 
         'F96-1': {}, 
         'F96': {}, 
         'TBLU': {},
         'variante': {}
     }
     for module in modules:
         if module in pdcrVariantes["large"]:
             flag_l = True
         if module in pdcrVariantes["medium"]:
             flag_m = True
         if module in pdcrVariantes["small"]:
             flag_s = True
         #print("Module i de la Lista: "+module)
         endpoint_Module= f"http://{host}:5000/api/get/{db}/modulos_fusibles/MODULO/=/{module}/_/=/_"
         #print("Endpoint del módulo"+endpoint_Module)
         response = requests.get(endpoint_Module).json()
         #print("Modulo Informacion",response)
         if "MODULO" in response:
             if len(response["MODULO"]) == 1: 
                 for j in response:
                     if j == "ID" or j == "MODULO":
                         response[j] = response[j][0]
                     else:
                         #print("j!!!!: ",j)
                         if j == "F96":
                             continue
                         response[j] = json.loads(response[j][0])
                         #print("response[j]",response[j])
                         for k in response[j]:
                             #print(k)
                             if response[j][k] != "empty":
                                 modularity[j][k] = [response[j][k],module]
     print("\t\t+++++++++++ FLAGS de",ILX,":+++++++++++\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l)
     if flag_l == True:
         variante = "PDC-R"
     if flag_m == True and flag_l == False:
         variante = "PDC-RMID"
     if flag_s == True and flag_m == False:
         variante = "PDC-RS"
     if flag_s == False and flag_m == False and flag_l == False:
         variante = "N/A"
         print("La caja no contiene módulos pertenecientes a las categorías.")
     modularity["variante"] = variante
     print("Variante de Caja: ",variante)
     return modularity

@app.route("/api/get/<db>/pdcr/variantes",methods=["GET"])
def variantesEvent(db):
    pdcrVariantes = {
    "small": [],
    "medium": [],
    "large": [],
    }

    endpoint = f"http://{host}:5000/api/get/{db}/definiciones/ACTIVE/=/1/_/_/_"
    pdcrVariantesDB = requests.get(endpoint).json()
    if "exception" in pdcrVariantesDB:
        endpoint = f"http://{host}:5000/api/get/{db}/definiciones/ACTIVO/=/1/_/_/_"
        pdcrVariantesDB = requests.get(endpoint).json()

    try:
        if len(pdcrVariantesDB["MODULO"]) > 0:
            #print("Cantidad de Módulos: ",len(pdcrVariantesDB["MODULO"]))
            #print("Lista de Módulos: ",pdcrVariantesDB["MODULO"])
            #print("Lista de Variantes: ",pdcrVariantesDB["VARIANTE"])
            for i in pdcrVariantesDB["MODULO"]:
                #print("Modulo Actual (i)",i)
                #print("Index de Modulo Actual (i)",pdcrVariantesDB["MODULO"].index(i))
                #print("Variante correspondiente a Modulo Actual: ",pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)])
                if pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-R":
                    pdcrVariantes["large"].append(i)
                    #print("ES UNA PDC-R LARGE")
                elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RMID":
                    #print("ES UNA PDC-R MEDIUM")
                    pdcrVariantes["medium"].append(i)
                elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RS":
                    #print("ES UNA PDC-R SMALL")
                    pdcrVariantes["small"].append(i)
    except Exception as ex:
        print("Variantes Exception: ", ex)
        return {"exception": ex.args}
    return pdcrVariantes

@app.route("/api/delete/<db>/<table>/<int:ID>",methods=["POST"])
def deleteEvent(table, ID,db):
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = db)
    except Exception as ex:
        print("delete connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(f"DELETE FROM {table} WHERE ID={ID}")
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("dele Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route('/database/<db>/<table>/<column_of_table_1>/<operation_1>/<val_1>/<column_of_table_2>/<operation_2>/<val_2>',methods=['GET'])
def value_of_a_tableEvent(table,column_of_table_1,operation_1,val_1,column_of_table_2,operation_2,val_2,db):
    if column_of_table_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if val_2=='_':
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'";'
        else:
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'" AND '+column_of_table_2+operation_2 +'"'+val_2+'";'
    print(query)
    #conexion con base de datos
    conexion =  pymysql.connect(host = host, user = user, passwd = password, database = db)
    cursor = conexion.cursor()
    cursor.execute(query)
    result = cursor.fetchone()
    
    if result == None:
        resp='NO HAY INFORMACION'
        response=resp
    else:
        resp='SI HAY INFORMACION'
        query = 'SELECT COLUMN_NAME FROM Information_Schema.Columns WHERE TABLE_NAME = ' + '"' + table + '";'
        cursor.execute(query)
        name_columns=cursor.fetchall()
        print(type(result))
        print(len(result))
        print(result)
        print(type(name_columns))
        print(len(name_columns))
        print(name_columns)

        dic={}
        for i in range(len(result)):
            dic[name_columns[i][0]]=result[i]
        print(dic)
        response=dic
    return response

################################################## Update Fijikura Server  ####################################################
@app.route("/seghm/get/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def famx2GET(table, column_1, operation_1, value_1, column_2, operation_2, value_2):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("Conexión Éxitosa")
    except Exception as ex:
        print("Conexión a P2 Exception: ", ex)
        return {"exception": ex.args}

    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            #result = cursor.fetchall()

            records = cursor.fetchall()
            insertObject = []
            columnNames = [column[0]
               for column in cursor.description
            ]

            for record in records:
               insertObject.append(dict(zip(columnNames, record)))
               #print("insertObject FINAL: ",insertObject)
            if len(insertObject) == 1:
                response = insertObject[0]
            elif len(insertObject) > 1:
                response = {}
                keys = list(insertObject[0])
                for key in keys:
                    response[key] = []
                    for item in insertObject:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": 0}
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/seghm/update/<table>/<int:ID>",methods=["POST"])
def famx2update(table, ID):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    flag_torque = False
    flag_vision = False
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("|||| SERVICIO UPDATE Conexión Éxitosa")
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "UPDATE " + table + f" SET"
        valores = []
        for key in data:
            try:
                if key == "DATETIME":
                    if data[key] == "AUTO":
                        data[key] = datetime.now().isoformat()
                if type(data[key]) == dict:
                    data[key] = json.dumps(data[key])
                valores.append(data[key])
            except Exception as ex:
                print("keys inspection Exception: ", ex)
            query += f' {key}= ?,'
            #print("primer Query: ",query)
            #print("Valores Final: ",valores)
        query = query[:-1]
        #print("query: ",query)
        query += f" WHERE ID={ID}"
        #print("query con += : ",query)
        with connection.cursor() as cursor:
            #print("dentro de cursor")
            items = cursor.execute(query, valores)
        connection.commit()
        response = {"items": 1}
    except Exception as ex:
        print("update Exception: ", ex)
        response = {"exception": 0}
    finally:
        connection.close()
        return response

@app.route('/contar/<table>/<column>', methods=['GET'])
def data_count(table, column):
    turnos = request.get_json(force=True)
    turnos = {
            "1":["07-00","16-59"],
            "2":["17-00","06-59"],
            }

    print("turnos:",turnos)
    ####### REVISAR EN QUÉ TURNO ESTÁ LA HORA ACTUAL
    for elemento in turnos:

        hora_iniciostr = turnos[elemento][0]
        hora_finstr = turnos[elemento][1]

        inicio_split = hora_iniciostr.split("-")
        hora_inicio = int(inicio_split[0])
        minuto_inicio = int(inicio_split[1])
            
        fin_split = hora_finstr.split("-")
        hora_fin = int(fin_split[0])
        minuto_fin = int(fin_split[1])

            
        #Se obtiene la Hora actual (int)
        horaActual = datetime.now().hour
        #Minutos Actuales
        minActual = datetime.now().minute
            

        #se detecta el tipo de jornada ....
        caso1 = False #inicio menor que fin (jornada normal)
        caso2 = False #fin menor que inicio (jornada con cambio de día)
            
        if hora_inicio < hora_fin:
            #print("hora inicio menor que hora fin")
            caso1 = True
        else:
            if hora_inicio == hora_fin:
                #print("hora inicio = a hora fin")
                if minuto_inicio < minuto_fin:
                    #print("minuto inicio menor que minuto fin")
                    caso1 = True
                else:
                    #print("minuto inicio mayor que minuto fin")
                    caso2 = True
            else:
                #print("hora inciio menor que hora fin")
                caso2 = True


        #Fecha Actual
        fechaActual = datetime.today()
        ##Segundos Actuales
        ##secActual = datetime.now().second
        #delta time de un día
        td = timedelta(days = 1)
        ayerfechaActual = fechaActual - td
        mañanafechaActual = fechaActual + td

        hoy_year =  datetime.now().year
        hoy_month = datetime.now().month
        hoy_day =   datetime.now().day

        ayer_year =  ayerfechaActual.year
        ayer_month = ayerfechaActual.month
        ayer_day =   ayerfechaActual.day

        mañana_year =  mañanafechaActual.year
        mañana_month = mañanafechaActual.month
        mañana_day =   mañanafechaActual.day

        inicio_query = ""
        fin_query = ""

        #AQUÍ YA SE SABE EL TIPO DE HORARIO QUE SE ESTÁ REVISANDO, 
        #HAY QUE VER SI LA HORA ACTUAL ESTÁ DENTRO DE ESTE HORARIO

        if caso1 == True:

            init_date = datetime(hoy_year, hoy_month, hoy_day, hora_inicio, minuto_inicio )
            end_date  = datetime(hoy_year, hoy_month, hoy_day, hora_fin,    minuto_fin )

            if init_date <= fechaActual <= end_date:
                inicio_query = str(init_date.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date.strftime('%Y-%m-%d-%H-%M'))
                break

        if caso2 == True:

            init_date1 = datetime(    hoy_year,     hoy_month,     hoy_day,  hora_inicio,  minuto_inicio )
            end_date1 =  datetime( mañana_year,  mañana_month,  mañana_day,  hora_fin,     minuto_fin )
                
            init_date2 = datetime(ayer_year, ayer_month, ayer_day, hora_inicio, minuto_inicio )
            end_date2 =  datetime( hoy_year,  hoy_month,  hoy_day,    hora_fin,    minuto_fin )

            if init_date1 <= fechaActual <= end_date1:
                inicio_query = str(init_date1.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date1.strftime('%Y-%m-%d-%H-%M'))
                break

            if init_date2 <= fechaActual <= end_date2:
                inicio_query = str(init_date2.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date2.strftime('%Y-%m-%d-%H-%M'))
                break

    ####### CONTAR LOS ARNESES QUE HAY o HA HABIDO ENTRE TAL FECHA Y TAL FECHA DEPENDIENDO DEL CASO

    print("--------------------inicio_query: ",inicio_query)
    print("--------------------   fin_query: ",fin_query)

    query= "SELECT * FROM " +table+" WHERE "+ column + ">=" + "'" + inicio_query + "' AND " + column + "<=" + "'" + fin_query + "';"
    print("query: ",query)

    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    
    except Exception as ex:
        print("data_count connection Exception: ", ex)
        response = {"conteo" : 0}
        return response


    try:
        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall() 
        pedidos = []

        for i in result: ##Buscando diferentes Valores en el rango de fecha
            indice = result.index(i)
            if result[indice]["RESULTADO"] == "BUENO" : ##Revisando si existen resets
                pedidos.append(result[indice]["RESULTADO"])
                #print(pedidos)
        mylist = list(dict.fromkeys(pedidos)) ## Eliminando valores duplicados
        #print(len(pedidos), 'The Big ONE')
        ###CONTAR LOS ITEMS LEÍDOS

        #if inicio_query == "":
        #    conteo = 0
        #else:
        #    if len(result):
        #        if isinstance(result, str):
        #            conteo = 1
        #        if isinstance(result, list):
        #            conteo = len(result)
        #            #print(result)
        #    else:
        #        conteo = 0

        response = {"conteo" : len(pedidos)}

    except Exception as ex:
        print("data_count cursor Exception: ", ex)
        response = {"conteo" :0}
        return response

    finally:
        connection.close()
        return response

@app.route('/horaxhora/<table>/<column>', methods=['GET'])
def horaxhora(table, column):
    turnos = request.get_json(force=True)
    turnos = {
            "1":["07-00","16-59"],
            "2":["17-00","06-59"],
            }

    print("turnos:",turnos)
    ####### REVISAR EN QUÉ TURNO ESTÁ LA HORA ACTUAL
    for elemento in turnos:

        hora_iniciostr = turnos[elemento][0]
        hora_finstr = turnos[elemento][1]

        inicio_split = hora_iniciostr.split("-")
        hora_inicio = int(inicio_split[0])
        minuto_inicio = int(inicio_split[1])
            
        fin_split = hora_finstr.split("-")
        hora_fin = int(fin_split[0])
        minuto_fin = int(fin_split[1])

            
        #Se obtiene la Hora actual (int)
        horaActual = datetime.now().hour
        #Minutos Actuales
        minActual = datetime.now().minute
            

        #se detecta el tipo de jornada ....
        caso1 = False #inicio menor que fin (jornada normal)
        caso2 = False #fin menor que inicio (jornada con cambio de día)
            
        if hora_inicio < hora_fin:
            #print("hora inicio menor que hora fin")
            caso1 = True
        else:
            if hora_inicio == hora_fin:
                #print("hora inicio = a hora fin")
                if minuto_inicio < minuto_fin:
                    #print("minuto inicio menor que minuto fin")
                    caso1 = True
                else:
                    #print("minuto inicio mayor que minuto fin")
                    caso2 = True
            else:
                #print("hora inciio menor que hora fin")
                caso2 = True


        #Fecha Actual
        fechaActual = datetime.today()
        ##Segundos Actuales
        ##secActual = datetime.now().second
        #delta time de un día
        td = timedelta(days = 1)
        ayerfechaActual = fechaActual - td
        mañanafechaActual = fechaActual + td

        hoy_year =  datetime.now().year
        hoy_month = datetime.now().month
        hoy_day =   datetime.now().day

        ayer_year =  ayerfechaActual.year
        ayer_month = ayerfechaActual.month
        ayer_day =   ayerfechaActual.day

        mañana_year =  mañanafechaActual.year
        mañana_month = mañanafechaActual.month
        mañana_day =   mañanafechaActual.day

        inicio_query = ""
        fin_query = ""

        #AQUÍ YA SE SABE EL TIPO DE HORARIO QUE SE ESTÁ REVISANDO, 
        #HAY QUE VER SI LA HORA ACTUAL ESTÁ DENTRO DE ESTE HORARIO

        if caso1 == True:

            init_date = datetime(hoy_year, hoy_month, hoy_day, hora_inicio, minuto_inicio )
            end_date  = datetime(hoy_year, hoy_month, hoy_day, hora_fin,    minuto_fin )

            if init_date <= fechaActual <= end_date:
                inicio_query = str(init_date.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date.strftime('%Y-%m-%d-%H-%M'))
                break

        if caso2 == True:

            init_date1 = datetime(    hoy_year,     hoy_month,     hoy_day,  hora_inicio,  minuto_inicio )
            end_date1 =  datetime( mañana_year,  mañana_month,  mañana_day,  hora_fin,     minuto_fin )
                
            init_date2 = datetime(ayer_year, ayer_month, ayer_day, hora_inicio, minuto_inicio )
            end_date2 =  datetime( hoy_year,  hoy_month,  hoy_day,    hora_fin,    minuto_fin )

            if init_date1 <= fechaActual <= end_date1:
                inicio_query = str(init_date1.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date1.strftime('%Y-%m-%d-%H-%M'))
                break

            if init_date2 <= fechaActual <= end_date2:
                inicio_query = str(init_date2.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date2.strftime('%Y-%m-%d-%H-%M'))
                break

    ####### CONTAR LOS ARNESES QUE HAY o HA HABIDO ENTRE TAL FECHA Y TAL FECHA DEPENDIENDO DEL CASO

    print("--------------------inicio_query: ",inicio_query)
    print("--------------------   fin_query: ",fin_query)

    query= "SELECT HM,RESULTADO,INICIO,FIN,USUARIO FROM " +table+" WHERE "+ column + ">=" + "'" + inicio_query + "' AND " + column + "<=" + "'" + fin_query + "';"
    print("query: ",query)

    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    
    except Exception as ex:
        print("data_count connection Exception: ", ex)
        response = {'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        return response


    try:
        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall() 
        #print("result",result)
        if len(result) > 0:
            response = {}
            keys = list(result[0])
            for key in keys:
                response[key] = []
                for item in result:
                    response[key].append(item.pop(key))   
            response["columns"] = keys
        else:
            response ={'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        

    except Exception as ex:
        print("data_count cursor Exception: ", ex)
        response = {'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        return response

    finally:
        connection.close()
        return response
### Area de consulta de datos
@app.route('/descargar/<db>/<table>/<task>')
def descargar(db, table, task):
    query = 'SELECT * FROM ' +table+' WHERE '+task+';'
    print(query)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            #print("result: ",result)
            #print(result[0].keys())
            arreglo = []
            hourEnd = []
            h =  result[0]
            valores_fila = ','.join( str(valor) for valor in h)
            #print(valores_fila)
            li = list(valores_fila.split(","))
            #print(li)
            li.remove('FUSIBLES')
            li.remove('QR_FET')
            li.append('QR_FET')

            li.remove('QR_MAQUINA')
            li.append('QR_MAQUINA')

            #li.remove('ALTURA')
            li.remove('REINTENTOS')
            #li.remove('INTENTOS_T')
            #li.remove('SCRAP')
            #li.remove('SERIALES')
            #li.remove('NOTAS')
            #li.remove('ANGULO')
            #Capitalizando Titulos,   TITULO -> Titulo
            capt = []
            for l in li: 
                c = l.capitalize()
                capt.append(c)
                
            arreglo.append(capt)
            #arreglo.append(valores_fila)
            if len(result) > 0:
                # Procesar los resultados por fila
                for fila in result:
                    del fila['FUSIBLES']
                    del fila['REINTENTOS']
                    #del fila['QR_MAQUINA']
                    #del fila['INTENTOS_VA']
                    #del fila['INTENTOS_T']
                    #del fila['SCRAP']
                    #del fila['SERIALES']
                    #del fila['NOTAS']
                    #del fila['ANGULO']
                    
                    dato = []
                    if fila["HM"] != 'HM000000000003':
                        arreglo.append(dato)
                        #print(fila)
                        # fila es la lista de elementos en diccionario que tiene una fila
                        # li es la lista de titulos ordenados manualmente
                        for i in li:

                            #print(i)
                            if 'NOTAS' in i:
                                Notepad = json.loads(fila[i])
                                #print(Notepad['VISION'][0])
                                dato.append(Notepad['VISION'][0])
                            
                            else: 
                                #print(fila["FIN"])
                                # Define las dos fechas como cadenas de texto
                                hourEnd.append(fila['FIN'])
                                # Convierte las cadenas de texto en objetos datetime
                                #fecha1 = datetime.strptime(fecha1_str, "%d-%m-%Y %H:%M:%S")
                                #fecha2 = datetime.strptime(fecha2_str, "%d-%m-%Y %H:%M:%S")
                                #print(fila[i])
                                dato.append(fila[i])
                        
                    
                #print(arreglo)

            else:
                response = {"items": items}
                #print(response)

    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}


    #######################  REALIZANDO FORMATO EXCEL ################
    #######################                           ################

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Historial'

    # Crea una nueva hoja en el libro
    #sheet2 = workbook.create_sheet("Graficar") # insert at first position
    #sheet2 = workbook["Graficar"]

    sheet['A1'] = '_____'

    for j in arreglo:
        sheet.append(j)
    alineacion_izquierda = Alignment(horizontal='left')
    
    sheet.insert_cols(idx=6,amount=1)
    sheet['F2'] = 'Duracion (dias, horas, minutos, segundos)'
    
    sheet['J2'] = 'Tiempo Ciclo'
    sheet['K2'] = 'Tiempo Ciclo Horas'
    
    sheet.title = 'Historial'
        # Agrega valores a la nueva columna (no necesitas calcular la diferencia en Python)
    for i in range(3, sheet.max_row + 1):
        formula = f'= IFERROR(ABS(D{i}-E{i}), 0)'  # Suponiendo que "Inicio" está en la columna D y "Fin" en la columna E
        
        # Suponiendo que "Inicio" está en la columna D y "Fin" en la columna E
        formulaB = f'= CONCATENATE( (PRODUCT( DAY(F{i}),24) + HOUR(F{i}) ), ":", MINUTE(F{i}), ":", SECOND(F{i}))'  
        
        formulaC = f'=IFERROR(ABS($D{i}-$E{i}), 0)'  #Tiempo Muerto en Decimal de Horas
        formulaD = f'=IFERROR(ABS($D{i}-$E{i}) * 24, 0)'  #Tiempo Muerto en Decimal de Horas

        sheet.cell(row=i, column=6, value=formula)


        sheet.cell(row=i, column=10, value=formulaC)
        sheet.cell(row=i, column=11, value=formulaD)

    lastfila = get_column_letter(sheet.max_column)+str(sheet.max_row) 


# ####FORMATOS
    formato_hora = NamedStyle(name = 'formato_hora')
    formato_hora.number_format = 'hh:mm:ss'

    # Supongamos que deseas aplicar el formato a la columna A (por ejemplo, de la fila 2 a la fila 100)
    columna = sheet['F']
    for celda in columna[2:sheet.max_row]:  # Excluye la primera fila si tiene encabezados
             celda.number_format = 'dd hh:mm:ss'

    # Supongamos que deseas aplicar el formato a la columna A (por ejemplo, de la fila 2 a la fila 100)
    columna = sheet['J']
    for celda in columna[2:sheet.max_row]:  # Excluye la primera fila si tiene encabezados
            celda.number_format = 'dd hh:mm:ss'

# Iterar sobre todas las columnas y ajustar sus anchos
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell.alignment = alineacion_izquierda
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet.column_dimensions[column_letter].width = adjusted_width
    # sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['F'].width = 20
    # sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['E'].width = 19
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['N'].width = 13
    sheet.column_dimensions['O'].width = 15
    sheet.column_dimensions['P'].width = 15
    # Agregar título de la tabla de clientes
    sheet['A1'] = f'Fujikura Automotive México Piedras Negras "Insercion"'
    sheet.merge_cells('A1:D1')


    tab = Table(displayName="Table1", ref="A2:" + lastfila)

    # Agregando Estilos de tabla
    style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    sheet.add_table(tab)

    # Formulario para calcular por columnas diferentes tareas
    sheet['N2'] = 'Promedio'
    sheet['N3'] = "= TEXT(AVERAGE(F3:F"+str(sheet.max_row)+'), "hh:mm:ss")'
    sheet['O2'] = "No terminados"
    sheet['O3'] = '= COUNTIF(H3:H'+str(sheet.max_row)+',"="&"RESET")'


    tab2 = Table(displayName="Table2", ref="N2:O3")

    # Agregando Formato a la tabla
    formulas = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab2.tableStyleInfo = formulas

    sheet.add_table(tab2)

    # Establecer estilos de fuente y color
    first_table_font = Font(color="124B43")  # Azul Marino
    second_table_font = Font(color="0043BB")  # Un tono más claro de rojo

    

    # Guardar el libro de Excel en un objeto en memoria
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo como respuesta para descarga
    return send_file(
        output,
        as_attachment=True,
        download_name='Fujikura Automotive México Piedras Negras.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
########################################################################################################################################


#@app.route("/seghm/post/<table>",methods=["POST"]) #MICROSERVICIO COMENTADO DEBIDO A QUE ACTUALMENTE NO SE UTILIZA PARA REALIZAR NINGÚN TIPO DE "POST" DE INFORMACIÓN A FAMX2, A DIFERENCIA DE TORQUE QUE SI PUBLICA EL HISTORIAL DE CADA REGISTRO
#def famx2POST(table):
#    def escape_name(s):
#        name = '{}'.format(s.replace('`', '``'))
#        return name
#    data = request.get_json(force=True)
#    #print("Data -*-*-*--*-*-*-**: ",data)
#    try:
#        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
#        print("|||| SERVICIO POST FAMX2 Conexión Éxitosa")
#    except Exception as ex:
#        print("famx2POST connection Exception: ", ex)
#        return {"exception": ex.args}
#    try:
#        query = "INSERT INTO " + table
#        keys = list(data)
#        #print("keys: ",keys)
#        cols = ', '.join(map(escape_name, keys))
#        placeholders = ', '.join(['?' for key in keys])
#        query += ' ({}) VALUES ({})'.format(cols, placeholders)
#        print("|||Query para POST: ",query)
#        #print("Data: ",data)
#        valores = []
#        for key in data:
#            try:
#                if key == "DATETIME":
#                    if data[key] == "AUTO":
#                        data[key] = datetime.now().isoformat()
#                if type(data[key]) == dict:
#                    data[key] = json.dumps(data[key])
#                valores.append(data[key])
#            except Exception as ex:
#                print("keys inspection Exception: ", ex)
#        with connection.cursor() as cursor:
#            items = cursor.execute(query, valores)
#        connection.commit()
#        response = {"items": 1} #Si el POST se realiza con éxito, al final regresará como respuesta el valor 1 asociado a la key "items"
#    except Exception as ex:
#        print("famx2POST Insert Exception: ", ex)
#        response = {"exception": ex.args}
#    finally:
#        connection.close()
#        return response
##################################################################################################

if __name__ == '__main__':
    """
        host: naapnx-famx2:8080
        user: amtc
        pass: amtc
    """
    app.run("0.0.0.0", 5000)
