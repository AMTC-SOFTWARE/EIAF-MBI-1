"""
        NOTAS:

for root, dirs, files in os.walk(dir_path):
    for file in files: 
  
        # change the extension from '.mp3' to 
        # the one of your choice.
        if file.endswith('.mp3'):
            print (root+'/'+str(file))

#Para la caja TBLU se usan fusibles ATO con color claro por lo que al color se le agrega un "_clear", por ejemplo "ATO,10,red_clear"

#"F400": "ATO,15,BLUE
"""
           
from copy import copy
import requests
import openpyxl
import json
import os
from model import model
import pymysql
import pyodbc
datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2=datos_conexion.datos_acceso()

modules = {}

fuses_types = {
    'PDC-P': {
        'MF1': "MULTI", 'MF2': "MULTI", 'F300': "ATO", 'F301': "MINI", 'F302': "MINI", 'F303': "MINI", 'F304': "MINI", 'F305': "MINI", 'F318': "MINI", 
        'F319': "MINI", 'F320': "MINI", 'F321': "MINI", 'F322': "MINI", 'F323': "MINI", 'F324': "MINI", 'F325': "MINI", 'F326': "ATO", 'F327': "ATO", 
        'F328': "ATO", 'F329': "ATO", 'F330': "ATO", 'F331': "ATO", 'F332': "ATO", 'F333': "ATO", 'F334': "ATO", 'F335': "ATO", 'E21': "CONN", 
        'E22': "CONN"
    },
    'PDC-D': {
        'F200': "MINI", 'F201': "MINI", 'F202': "MINI", 'F203': "MINI", 'F204': "MINI", 'F205': "MINI", 'F206': "MINI", 'F207': "MINI", 'F208': "MINI", 
        'F209': "ATO", 'F210': "ATO", 'F211': "ATO", 'F212': "ATO", 'F213': "ATO", 'F214': "ATO", 'F215': "ATO", 'F216': "ATO", 'F217': "MINI", 
        'F218': "MINI", 'F219': "MINI", 'F220': "MINI", 'F221': "MINI", 'F222': "MINI", 'F223': "MINI", 'F224': "MINI", 'F225': "MINI", 'F226': "MINI", 
        'F227': "MINI", 'F228': "MINI", 'F229': "MINI", 'F230': "MINI", 'F231': "MINI", 'F232': "MINI"
    },
    'PDC-R': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F449': "MAXI", 'F448': "MAXI", 'F447': "MAXI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 
        'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'F462': "MAXI", 
        'F463': "MAXI", 'F464': "MAXI", 'F465': "MINI", 'F466': "MINI", 'F467': "MINI", 'F468': "MINI", 'F469': "MINI", 'F470': "MINI", 'F471': "ATO", 
        'F472': "ATO", 'F473': "ATO", 'F474': "ATO", 'F475': "ATO", 'F476': "ATO", 'F477': "ATO", 'F478': "ATO", 'F479': "ATO", 'F480': "ATO", 
        'F481': "ATO", 'F482': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F96': "ATO"
    },
    'PDC-RMID': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-RS': {
        'F400': "ATO", 'F401': "ATO", 'F402': "ATO", 'F403': "ATO", 'F404': "ATO", 'F405': "ATO", 'F411': "MINI", 'F410': "MINI", 'F409': "MINI", 
        'F408': "MINI", 'F407': "MINI", 'F406': "MINI", 'F412': "ATO", 'F413': "ATO", 'F414': "ATO", 'F415': "ATO", 'F416': "ATO", 'F417': "ATO", 
        'F420': "MAXI", 'F419': "MAXI", 'F418': "MAXI", 'F421': "ATO", 'F422': "ATO", 'F423': "ATO", 'F424': "ATO", 'F425': "ATO", 'F426': "ATO", 
        'F427': "MINI", 'F428': "MINI", 'F429': "MINI", 'F430': "MINI", 'F431': "MINI", 'F437': "MINI", 'F438': "MINI", 'F439': "MINI", 'F440': "MINI", 
        'F441': "MINI", 'F432': "MINI", 'F433': "MINI", 'F434': "MINI", 'F435': "MINI", 'F436': "MINI", 'F442': "MINI", 'F443': "MINI", 'F444': "MINI", 
        'F445': "MINI", 'F446': "MINI", 'F450': "ATO", 'F451': "ATO", 'F452': "ATO", 'F453': "ATO", 'F454': "ATO", 'F455': "ATO", 'F456': "ATO", 
        'F457': "ATO", 'F458': "ATO", 'F459': "ATO", 'F460': "ATO", 'F461': "ATO", 'RELX': "RELAY", 'RELU': "RELAY", 'RELT': "RELAY", 'F449': "MAXI", 
        'F448': "MAXI", 'F447': "MAXI", 'F96': "ATO"
    },
    'PDC-S': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S9': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S21': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S17': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S19': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'PDC-S20': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'F96-1': {
        'F96': "ATO"
    },
    'F96': {
        'F96': "ATO"
    }, 
    'TBLU': {
        '9': "ATO", '8': "ATO", '7': "ATO", '6': "ATO", '5': "ATO", '4': "ATO", '3': "ATO", '2': "ATO", '1': "ATO"
    }
}

fuses_value = {
    'PDC-P': {
        'MF1': '', 'MF2': '', 'F300': '', 'F301': '', 'F302': '', 'F303': '', 'F304': '', 'F305': '', 'F318': '', 
        'F319': '', 'F320': '', 'F321': '', 'F322': '', 'F323': '', 'F324': '', 'F325': '', 'F326': '', 'F327': '', 
        'F328': '', 'F329': '', 'F330': '', 'F331': '', 'F332': '', 'F333': '', 'F334': '', 'F335': '', 'E21': '', 
        'E22': ''
    },
    'PDC-D': {
        'F200': '', 'F201': '', 'F202': '', 'F203': '', 'F204': '', 'F205': '', 'F206': '', 'F207': '', 'F208': '', 
        'F209': '', 'F210': '', 'F211': '', 'F212': '', 'F213': '', 'F214': '', 'F215': '', 'F216': '', 'F217': '', 
        'F218': '', 'F219': '', 'F220': '', 'F221': '', 'F222': '', 'F223': '', 'F224': '', 'F225': '', 'F226': '', 
        'F227': '', 'F228': '', 'F229': '', 'F230': '', 'F231': '', 'F232': ''
    },
    'PDC-R': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F449': '', 'F448': '', 'F447': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 
        'F454': '', 'F455': '', 'F456': '', 'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'F462': '', 
        'F463': '', 'F464': '', 'F465': '', 'F466': '', 'F467': '', 'F468': '', 'F469': '', 'F470': '', 'F471': '', 
        'F472': '', 'F473': '', 'F474': '', 'F475': '', 'F476': '', 'F477': '', 'F478': '', 'F479': '', 'F480': '', 
        'F481': '', 'F482': '', 'RELX': '', 'RELU': '', 'RELT': ''
    },
    'PDC-RMID': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 'F454': '', 'F455': '', 'F456': '', 
        'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'RELX': '', 'RELU': '', 'RELT': '', 'F449': '', 
        'F448': '', 'F447': '', 'F96': ''
    },
    'PDC-RS': {
        'F400': '', 'F401': '', 'F402': '', 'F403': '', 'F404': '', 'F405': '', 'F411': '', 'F410': '', 'F409': '', 
        'F408': '', 'F407': '', 'F406': '', 'F412': '', 'F413': '', 'F414': '', 'F415': '', 'F416': '', 'F417': '', 
        'F420': '', 'F419': '', 'F418': '', 'F421': '', 'F422': '', 'F423': '', 'F424': '', 'F425': '', 'F426': '', 
        'F427': '', 'F428': '', 'F429': '', 'F430': '', 'F431': '', 'F437': '', 'F438': '', 'F439': '', 'F440': '', 
        'F441': '', 'F432': '', 'F433': '', 'F434': '', 'F435': '', 'F436': '', 'F442': '', 'F443': '', 'F444': '', 
        'F445': '', 'F446': '', 'F450': '', 'F451': '', 'F452': '', 'F453': '', 'F454': '', 'F455': '', 'F456': '', 
        'F457': '', 'F458': '', 'F459': '', 'F460': '', 'F461': '', 'RELX': '', 'RELU': '', 'RELT': '', 'F449': '', 
        'F448': '', 'F447': ''
    },
    'F96': {'F96': ''},    
    'PDC-S': {
        '1': '', '2': '', '3': '', '4': '', '5': '', '6': ''
    },
    'PDC-S9': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S21': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S17': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    }, 
    'PDC-S19': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'PDC-S20': {
        '1': "ATO", '2': "ATO", '3': "ATO", '4': "ATO", '5': "ATO", '6': "ATO"
    },
    'F96-1': {
        'F96': "ATO"
    },
    'TBLU': {
        '9': '', '8': '', '7': '', '6': '', '5': '', '4': '', '3': '', '2': '', '1': ''
    }
}

fuses_color = {
    #"1":    "negro", HMTEST ILX296270B1031517 EL.
    "5":    {"N000000008698":"beige", "N000000008708":"beige", "N000000004202":"beigeClear", "N000000006465":"beige"},
    "7.5":  {"N000000008699":"brown", "N000000008709":"brown", "N000000006466":"brown"},
    "10":   {"N000000008700":"red", "N000000008710":"red", "N000000004204":"redClear"},
    "15":   {"N000000008701":"blue", "N000000008711":"blue"},
    "20":   {"N000000008702":"yellow"},
    "25":   {"N000000008703":"white"},
    "30":   {"N000000008704": "green", "N000000007658":"green"},
    "40":   {"N000000007659": "amber"},
    "50":   {"N000000007660":"red"},
    "60":   {"A0009821923":"red"},
    "70":   {"A0025429419":"gray"}
    # "60":   "azul"
    }

##################################### Modules management #################################
def makeModules(data):
    global modules
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    modules = {}
    print("#################### Modules ####################")
    print("Modulos anteriormente cargados: ",modules)
    dir_path = os.path.join(os.getcwd(), '..\\modules\\')
    file_name = None
    # Establece la conexión a la base de datos
    connection = pymysql.connect(
        host=host,
        user=user,
        password=password,
        database= data
    )
    cursor = connection.cursor()
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    if "Acomodos Modularidades" in sheet or "MFB" in sheet or "BATTERY" in sheet:
                        continue
                    currentSheet = file[sheet]
                    for column in range(8, currentSheet.max_column + 1):
                        module = currentSheet.cell(row = 3, column = column).value
                        if not(module in modules):
                            modules[module] = {}
                            #print("Modulo: ", module)
                        for row in range(5,currentSheet.max_row  + 1):
                            value = currentSheet.cell(row = row, column = column).value
                            if value == "x" or value == "X":
                                box = currentSheet.cell(row = row, column = 1).value
                                box = box.strip()
                                mercedes =  currentSheet.cell(row = row, column = 4).value
                                if box =="Fuse Box F55":
                                    box = "TBLU"
                                    #print("mercedes ", mercedes)
                                fuse = currentSheet.cell(row = row, column = 2).value
                                if box == "TBLU":
                                    fuse = fuse.replace("A", "")
                                #     #print("mercedes ", mercedes)
                                if "PDC-S" in  box or "F96" in box:
                                    
                                        #print('BOX', box)
                                        #print('BOX', data)
                                        try:
                                            # Consulta para verificar si la columna existe
                                            check_column_query = f"""
                                            SELECT COUNT(*)
                                            FROM INFORMATION_SCHEMA.COLUMNS 
                                            WHERE TABLE_NAME = 'modulos_fusibles' AND COLUMN_NAME = '{box}';
                                            """

                                            # Ejecutar la consulta de verificación
                                            cursor.execute(check_column_query)
                                            column_exists = cursor.fetchone()[0]

                                            # Si la columna no existe, agregarla
                                            if column_exists == 0:
                                                add_column_query = f"""
                                                ALTER TABLE `modulos_fusibles` 
                                                ADD COLUMN `{box}` LONGTEXT CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL;
                                                """
                                                cursor.execute(add_column_query)
                                                print(f"Columna '{box}' agregada a la tabla '{data}'.")
                                            #else:
                                                #print(f"La columna '{box}' ya existe en la tabla '{data}'.")

                                            # Confirmar los cambios
                                            connection.commit()

                                            fuse = str(fuse)
                                            fuse = fuse.strip()
                                        except Exception as ex :
                                            print('Exception BOX', ex)
                                        
                                #print("Tipo del Fuse Ya convertido: ",type(fuse))
                                if box == "PDC-R":
                                    if module in pdcrVariantes["large"]:
                                        box = "PDC-R"
                                    elif module in pdcrVariantes["medium"]:
                                        box = "PDC-RMID"
                                    elif module in pdcrVariantes["small"]:
                                        box = "PDC-RS"
                                    else:
                                        box = "PDC-RS"
                                    if fuse == "X" or fuse == "T" or fuse == "U":
                                        fuse = "REL" + fuse
                                amp = currentSheet.cell(row = row, column = 7).value
                                if not(box in modules[module]):
                                    modules[module][box] = {}
                                modules[module][box][fuse] = [amp[:-1], mercedes]
                                

                os.remove(root+'\\'+ file_name)

    structured_data = []
    for module in modules:
        temp = {
            "DBEVENT": data,
            "MODULO": "",
            "PDC-R": {},
            "PDC-RMID": {},
            "PDC-RS": {},
            "PDC-S": {},
            "TBLU": {},
            "PDC-D": {},
            "PDC-P": {},
            "F96": {},
            "F96-1": {},
            "PDC-S9": {}, 
            "PDC-S19": {},
            "PDC-S20": {},
            "PDC-S21": {}, 
            "PDC-S17": {}, 
            }

        temp["MODULO"] = module
        for box in modules[module]:
            for fuse in modules[module][box]:
                try:
                    Type    = fuses_types[box][fuse]
                    amp     = modules[module][box][fuse][0]
                    mercedes     = modules[module][box][fuse][1]
                    print('TYPE',Type)
                    print('AMP',amp)
                    # #print(fuses_types['PDC-S21']['3'])
                    color = ""
                    # if Type == "RELAY":
                    #      if amp == "60":
                    #          color = "red"
                    #      elif amp == "70":
                    #          color = "gray"
                    # else:
                    color   = fuses_color[amp][mercedes]
                    temp[box][fuse] = Type + "," + amp + "," + color
                except Exception as ex:
                    print("\nexception in [",module,"] [",box,"] [",fuse,"]")
                    print(ex)
        structured_data.append(temp)

    print ("\n total de modulos: ",len(structured_data))

    return structured_data

def updateModules(data):
    print("updating")
    tabla = data[0]["DBEVENT"]
    print("TABLAAAAA Update Modules+-+-+-+-: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/modulos_fusibles/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/modulos_fusibles"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modulos_fusibles/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def pdcrVariants (data):
    """

            IN CONSTRUCTION


    PDC-R small:  A2239060902
    PDC-R MEDIUM:  A2239061002
    PDC-R LARGE:  A2239061102
    """
    print("#################### pdcrVariants ####################")
    dir_path = os.path.join(os.getcwd(), '..\\FAAJISPREV\\')
    file_name = None
    rows = []
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if temp.endswith('.txt'):
                fic = open(dir_path + file_name)
                lines = list(fic)
                for i in lines:
                    i = i[:-1]
                    rows.append(i.split())
                print(len(lines))
                for i in range(5):
                    print(lines[i])

def refreshModules(data):
    data = makeModules(data)
    updateModules(data)

################################### Modularities management ##############################
def makeModularities(data):
    global modules
    print("#################### Modularities ####################")
    endpoint = f"http://{host}:5000/api/get/{data}/modulos_fusibles/all/-/-/-/-/-"
    modulesExisting = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos: ",modulesExisting["MODULO"])
    dir_path = os.path.join(os.getcwd(), '..\\ILX\\')
    file_name = None
    modularities = []
    modulosFaltantes = []
    ilxfaltantes = {
        "ILX": {},
        "Modulos": {}
        }
    flujo = ""
    numero = ""
    if 'izquierda' in data:
        print('EVENTO DE CONDUCCION IZQUIERDA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'ILZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'ILX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'ILX'
            numero = '294'
    if 'derecha' in data:
        print('EVENTO DE CONDUCCION DERECHA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'IRZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'IRX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'IRX'
            numero = '294'        
    flujo_numero = flujo + numero

    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
        
            if not(flujo_numero in file_name):# SI NO se encuentra el nombre esperado de inicio para un arnés de este tipo:
                ilxfaltantes["ILX"][ILX] = []
                ilxfaltantes["ILX"][ILX].append("No es un DAT válido para este evento") #se agrega el mensaje que no es un DAT válido
                #ilxfaltantes["ILX"][ILX]["torque"].append("No es un DAT válido para este evento") #se agrega el mensaje que no es un DAT válido
                modulosFaltantes.append(ILX) #se agrega a la lista final de módulos faltantes para que aparezca en pantalla
                ilxfaltantes["Modulos"] = modulosFaltantes #se actualiza esta lista
                os.remove(root+'\\'+ file_name) #se elimina el archivo de los DATS
            else:
                if temp.endswith('.dat'):
                    fic = open(dir_path + file_name)
                    lines = list(fic)
                    csv = ""
                for line in lines:
                    csv += line.rsplit(sep = "=")[-1][:-1] + ","
                csv = csv[:-1]
                fic.close()
                temp = {
                    "DBEVENT": data,
                    "MODULARIDAD": ILX,
                    "FECHA": "AUTO",
                    "MODULOS_FUSIBLES": csv,
                    "ACTIVO": 1
                    }
                #print("ILX: ",ILX)
                #print("Modulos que tiene: ",csv)
                #print("Modulos que tiene TIPO: ",type(csv))
                #print("Modulos que tiene el ILX: ",csv.split(","))
                #print("Modulos que tiene convertido a array TIPO: ",type(csv.split(",")))
                modulosDesconocidos = set(csv.split(",")) - set(modulesExisting["MODULO"])
                #print("Comparación; Modulos del ILX que NO están en la base de datos: ", modulosDesconocidos)
                #print("Comparación; ILXFALTANTES: ", ilxfaltantes)
                #print("Comparación; Modulos del ILX que NO están en la base de datos LEN: ", len(modulosDesconocidos))
                #print("Comparación tipo", type(modulosDesconocidos))
                if len(modulosDesconocidos) == 0:
                    modularities.append(temp)
                else:
                    ilxfaltantes["ILX"][ILX] = []
                    for e in modulosDesconocidos:
                        ilxfaltantes["ILX"][ILX].append(e)
                        #print(e, "AAAAAAAA")
                        if not(e in modulosFaltantes):
                            modulosFaltantes.append(e)
                    ilxfaltantes["Modulos"] = modulosFaltantes
                os.remove(root+'\\'+ file_name)

        #print("MODULOS FALTANTES FINAL : ",modulosFaltantes)
        #print("ILX que NO se cargaron a la estación : ",ilxfaltantes)
    if len(modularities) != 0:
        updateModularities(modularities)
    return ilxfaltantes
 
def updateModularities(data):
    print("updating")
    print("Data dentro de Upload Modularities: ",data)
    tabla = data[0]["DBEVENT"]
    endpoint = f"http://{host}:5000/api/get/{tabla}/modularidades/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULARIDAD" in existing):
        existing["MODULARIDAD"] = []
    for i in data:
        try:
            if not(i["MODULARIDAD"] in existing["MODULARIDAD"]):
                endpoint = f"http://{host}:5000/api/post/modularidades"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULARIDAD"].index(i["MODULARIDAD"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modularidades/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

##################################### Determinantes management #################################
def refreshDeterminantes(data,usuario):
    data = makeDeterminantes(data,usuario)
    updateDeterminantes(data)

def makeDeterminantes(data,usuario):
    global determinantes
    determinantes = {
        "PDC-RS":[],
        "PDC-RMID":[],
        "PDC-R":[]
        }
    print("#################### Modulos Determinantes ####################")
    print("Modulos anteriormente cargados: ",determinantes)
    print("DATA que se pasa como arugmento a Determinantes",data)
    print("USUARIO que se pasa como arugmento a Determinantes",usuario)
    dir_path = os.path.join(os.getcwd(), '..\\determinantes\\')
    file_name = None
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    currentSheet = file[sheet]
                    columnas = ["PDC-RS","PDC-RMID","PDC-R"]
                    for variante in columnas:
                        if variante == "PDC-RS":
                            col = 2
                        if variante == "PDC-RMID":
                            col = 5
                        if variante == "PDC-R":
                            col = 8
                        #print("Col Actual: ",col)
                        for row in range(3, currentSheet.max_row + 1):
                            module = currentSheet.cell(column = col, row = row).value
                            if not(module in determinantes[variante]):
                                if module != None:
                                    #print("FILA: ", row, " COLUMNA: ", col)
                                    determinantes[variante].append(module)
                                    #print("Modulo: ", module)
                print("Arreglo final de determinantes: ",determinantes)
                os.remove(root+'\\'+ file_name)

    structured_data = []
    for variante in determinantes:
        print("Variante: ",variante)
        for module in determinantes[variante]:
            print("Modulo: ",module)
            temp = {
            "DBEVENT": data,
            "MODULO": module,
            "VARIANTE": variante,
            "DATETIME": "AUTO",
            "USUARIO": usuario,
            "ACTIVO": 1
            }
            structured_data.append(temp)

    print ("\n total de modulos: ",len(structured_data))

    return structured_data

def updateDeterminantes(data):
    print("updating")
    tabla = data[0]["DBEVENT"]
    print("Update determinantes evento+-+-+-+-: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/definiciones/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/definiciones"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/definiciones/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)



if __name__ == '__main__':
    print("finished")
    #refreshModules()
    #data = makeModules()
    #updateModules(data)
    #pdcrVariants("dumie")